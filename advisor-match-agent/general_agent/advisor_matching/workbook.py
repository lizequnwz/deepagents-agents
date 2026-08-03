"""Safe, styled advisor-match workbook projection and verification."""

from __future__ import annotations

import json
from collections.abc import Callable, Iterable, Mapping
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from general_agent.advisor_matching.input_loader import input_mapping_fingerprint
from general_agent.advisor_matching.schemas import (
    InputMapping,
    InputSummary,
    MatchCandidate,
    MatchCounts,
    MatchDecision,
    ReferenceSnapshotManifest,
)

SHEETS = ("Matched", "Review Required", "Original Input", "Run Summary")
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF")
_BODY_FONT = Font(name="Calibri", size=11)
_ALT_FILL = PatternFill("solid", fgColor="F4F7FA")
_STATUS_FILL = {
    "Matched": PatternFill("solid", fgColor="D9EAD3"),
    "Ambiguous Match": PatternFill("solid", fgColor="FFF2CC"),
    "No Match": PatternFill("solid", fgColor="F4CCCC"),
}
_HIDDEN_AUDIT_HEADERS = {
    "Review Item ID",
    "Rule ID",
    "Confidence",
    "Automated Status",
    "Decision Source",
    "Duplicate Group",
}


def write_match_workbook(
    output: Path,
    *,
    session_id: str,
    decisions: list[MatchDecision],
    counts: MatchCounts,
    mapping: InputMapping,
    input_summary: InputSummary,
    source_name: str,
    source_sha256: str,
    reference: ReferenceSnapshotManifest,
    policy_version: str,
    session_status: str = "Reviewing",
    session_revision: int = 1,
    source_transformation: Mapping[str, object] | None = None,
) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = output.with_name(f".{output.stem}.building{output.suffix}")
    workbook = Workbook(write_only=True)
    _write_matched(workbook.create_sheet("Matched"), decisions)
    _write_review(workbook.create_sheet("Review Required"), decisions)
    _write_original(workbook.create_sheet("Original Input"), decisions)
    _write_summary(
        workbook.create_sheet("Run Summary"),
        session_id=session_id,
        decisions=decisions,
        counts=counts,
        mapping=mapping,
        input_summary=input_summary,
        source_name=source_name,
        source_sha256=source_sha256,
        reference=reference,
        policy_version=policy_version,
        session_status=session_status,
        session_revision=session_revision,
        source_transformation=source_transformation,
    )
    workbook.save(temporary)
    verify_match_workbook(
        temporary,
        expected_rows=len(decisions),
        expected_counts=counts,
    )
    temporary.replace(output)


def verify_match_workbook(
    path: Path,
    *,
    expected_rows: int | None = None,
    expected_counts: MatchCounts | None = None,
) -> dict[str, int]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if tuple(workbook.sheetnames) != SHEETS:
            raise ValueError(f"Unexpected workbook sheets: {workbook.sheetnames!r}.")
        matched = sum(
            1 for _ in workbook["Matched"].iter_rows(min_row=2, values_only=True)
        )
        review_sheet = workbook["Review Required"]
        review_headers = [cell.value for cell in next(review_sheet.iter_rows(max_row=1))]
        try:
            review_id_index = review_headers.index("Review Item ID")
            review_status_index = review_headers.index("Status")
        except ValueError as exc:
            raise ValueError(
                "Review Required is missing its audit item ID or status."
            ) from exc
        review_statuses: dict[str, str] = {}
        for row in review_sheet.iter_rows(min_row=2, values_only=True):
            if len(row) <= max(review_id_index, review_status_index):
                continue
            review_id = row[review_id_index]
            if not review_id:
                continue
            status = row[review_status_index]
            prior = review_statuses.setdefault(str(review_id), str(status))
            if prior != status:
                raise ValueError(
                    f"Review item {review_id!r} has inconsistent statuses."
                )
        review_item_ids = set(review_statuses)
        original = sum(
            1
            for _ in workbook["Original Input"].iter_rows(
                min_row=2, values_only=True
            )
        )
        if expected_rows is not None and original != expected_rows:
            raise ValueError(
                f"Original Input has {original} rows; expected {expected_rows}."
            )
        if matched + len(review_item_ids) != original:
            raise ValueError(
                "Matched and Review Required decisions do not reconcile to "
                "Original Input."
            )
        ambiguous = sum(
            status == "Ambiguous Match" for status in review_statuses.values()
        )
        no_match = sum(status == "No Match" for status in review_statuses.values())
        if ambiguous + no_match != len(review_item_ids):
            raise ValueError("Review Required contains an invalid decision status.")
        actual_counts = MatchCounts(
            matched=matched,
            ambiguous_match=ambiguous,
            no_match=no_match,
        )
        if expected_counts is not None and actual_counts != expected_counts:
            raise ValueError(
                "Workbook status counts do not match the persisted decisions."
            )
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        raise ValueError(
                            f"Formula cell is not allowed: "
                            f"{sheet.title}!{cell.coordinate}."
                        )
        return {
            "matched": matched,
            "ambiguous_match": ambiguous,
            "no_match": no_match,
            "review_items": len(review_item_ids),
            "original": original,
        }
    finally:
        workbook.close()


def _write_matched(sheet, decisions: list[MatchDecision]) -> None:
    headers = [
        "Source Row",
        "Status",
        "Input Name",
        "Input Firm",
        "Input Location",
        "Input CRD",
        "Input Email",
        "Advisor CRD",
        "Advisor Name",
        "Advisor Firm",
        "Advisor Location",
        "Advisor Email",
        "Reason",
        "Warnings",
        "Supporting Evidence",
        "Conflicting Evidence",
        "Context",
        "Review Item ID",
        "Rule ID",
        "Confidence",
        "Automated Status",
        "Decision Source",
        "Duplicate Group",
    ]

    def rows() -> Iterable[list[object]]:
        for decision in decisions:
            if decision.status == "Matched":
                yield _matched_row(decision)

    _write_table(sheet, headers, rows, hidden=_HIDDEN_AUDIT_HEADERS)


def _write_review(sheet, decisions: list[MatchDecision]) -> None:
    headers = [
        "Source Row",
        "Status",
        "Reason",
        "Input Name",
        "Input Firm",
        "Input Location",
        "Input CRD",
        "Input Email",
        "Candidate Rank",
        "Candidate CRD",
        "Candidate Name",
        "Candidate Firm",
        "Candidate Location",
        "Candidate Email",
        "Supporting Evidence",
        "Conflicting Evidence",
        "Context",
        "Warnings",
        "Review Item ID",
        "Rule ID",
        "Confidence",
        "Automated Status",
        "Decision Source",
        "Duplicate Group",
    ]

    def rows() -> Iterable[list[object]]:
        for decision in decisions:
            if decision.status == "Matched":
                continue
            candidates: list[MatchCandidate | None] = decision.candidates or [None]
            for rank, candidate in enumerate(candidates, start=1):
                yield _review_row(decision, candidate, rank=rank)

    _write_table(sheet, headers, rows, hidden=_HIDDEN_AUDIT_HEADERS)


def _write_original(sheet, decisions: list[MatchDecision]) -> None:
    source_headers: list[str] = []
    for decision in decisions:
        for header in decision.source_values:
            if header not in source_headers:
                source_headers.append(header)
    headers = ["Source Row", *source_headers]

    def rows() -> Iterable[list[object]]:
        for decision in decisions:
            yield [
                decision.source_row_number,
                *[
                    decision.source_values.get(header, "")
                    for header in source_headers
                ],
            ]

    _write_table(sheet, headers, rows, hidden=set())


def _write_summary(
    sheet,
    *,
    session_id: str,
    decisions: list[MatchDecision],
    counts: MatchCounts,
    mapping: InputMapping,
    input_summary: InputSummary,
    source_name: str,
    source_sha256: str,
    reference: ReferenceSnapshotManifest,
    policy_version: str,
    session_status: str,
    session_revision: int,
    source_transformation: Mapping[str, object] | None,
) -> None:
    mapping_fingerprint = input_mapping_fingerprint(
        source_sha256=source_sha256, mapping=mapping
    )
    values: list[tuple[str, object]] = [
        ("Session ID", session_id),
        ("Session Status", session_status),
        ("Session Revision", session_revision),
        ("Source File", source_name),
        ("Selected Worksheet", mapping.sheet_name or "First worksheet / CSV"),
        ("Header Mode", "Headerless" if mapping.header_row is None else "Headed"),
        ("Header Row", mapping.header_row if mapping.header_row is not None else ""),
        ("Data Rows", input_summary.data_row_count),
        ("Skipped Blank Rows", input_summary.blank_row_count),
        ("Skipped Preamble Rows", input_summary.preamble_row_count),
        ("Firm Column Missing", input_summary.firm_column_missing),
        ("Rows Missing Firm and Strong IDs", input_summary.missing_firm_row_count),
        ("Matched", counts.matched),
        ("Ambiguous Match", counts.ambiguous_match),
        ("No Match", counts.no_match),
        ("Input Mapping", json.dumps(mapping.model_dump(mode="json"), sort_keys=True)),
        ("Mapping Fingerprint", mapping_fingerprint),
        ("Source SHA-256", source_sha256),
        ("Reference Snapshot ID", reference.reference_snapshot_id),
        ("Reference Rows", reference.row_count),
        ("Reference SHA-256", reference.sha256),
        ("Reference Retrieved At", reference.retrieved_at.isoformat()),
        ("Policy Version", policy_version),
        ("Rows With Warnings", sum(bool(item.warnings) for item in decisions)),
        (
            "Duplicate Rows",
            sum(item.duplicate_group is not None for item in decisions),
        ),
        ("Generated At", datetime.now(UTC).isoformat()),
    ]
    if source_transformation:
        values.extend(
            [
                ("Input Transformation", "User-confirmed bulk firm augmentation"),
                (
                    "Derived From Attachment ID",
                    source_transformation.get("source_attachment_id", ""),
                ),
                (
                    "Original Source SHA-256",
                    source_transformation.get("source_sha256", ""),
                ),
                (
                    "User-Supplied Firm",
                    source_transformation.get("firm_name", ""),
                ),
                (
                    "Rows Augmented With Firm",
                    source_transformation.get("rows_updated", ""),
                ),
            ]
        )
    for field, binding in mapping.field_bindings().items():
        values.append(
            (
                f"Mapped {field.replace('_', ' ').title()}",
                ", ".join(
                    f"index {reference.index} / {reference.header!r}"
                    for reference in binding.columns
                ),
            )
        )

    def rows() -> Iterable[list[object]]:
        for field, value in values:
            yield [field, value]

    _write_table(sheet, ["Field", "Value"], rows, hidden=set())


def _matched_row(decision: MatchDecision) -> list[object]:
    advisor = decision.matched_advisor
    return [
        decision.source_row_number,
        decision.status,
        _input_name(decision),
        decision.mapped_values.get("firm_name", ""),
        _location(decision.mapped_values),
        decision.mapped_values.get("crd_number", ""),
        decision.mapped_values.get("email", ""),
        advisor.crd_number if advisor else "",
        _advisor_name(advisor),
        advisor.firm_name if advisor else "",
        _advisor_location(advisor),
        advisor.email if advisor else "",
        decision.explanation,
        "; ".join(decision.warnings),
        "; ".join(advisor.supporting_evidence) if advisor else "",
        "; ".join(advisor.conflicting_evidence) if advisor else "",
        "; ".join(advisor.contextual_evidence) if advisor else "",
        decision.review_item_id,
        decision.rule_id,
        decision.confidence,
        decision.automated_status or decision.status,
        decision.decision_source,
        decision.duplicate_group or "",
    ]


def _review_row(
    decision: MatchDecision,
    candidate: MatchCandidate | None,
    *,
    rank: int,
) -> list[object]:
    return [
        decision.source_row_number,
        decision.status,
        decision.explanation,
        _input_name(decision),
        decision.mapped_values.get("firm_name", ""),
        _location(decision.mapped_values),
        decision.mapped_values.get("crd_number", ""),
        decision.mapped_values.get("email", ""),
        rank if candidate else "",
        candidate.crd_number if candidate else "",
        _advisor_name(candidate),
        candidate.firm_name if candidate else "",
        _advisor_location(candidate),
        candidate.email if candidate else "",
        "; ".join(candidate.supporting_evidence) if candidate else "",
        "; ".join(candidate.conflicting_evidence) if candidate else "",
        "; ".join(candidate.contextual_evidence) if candidate else "",
        "; ".join(decision.warnings),
        decision.review_item_id,
        decision.rule_id,
        decision.confidence,
        decision.automated_status or decision.status,
        decision.decision_source,
        decision.duplicate_group or "",
    ]


def _write_table(
    sheet,
    headers: list[str],
    rows_factory: Callable[[], Iterable[list[object]]],
    *,
    hidden: set[str],
) -> None:
    widths = [max(10, len(header) + 2) for header in headers]
    row_count = 0
    for values in rows_factory():
        row_count += 1
        for index, value in enumerate(values):
            widths[index] = max(
                len(headers[index]) + 2,
                min(60, max(widths[index], _display_width(value) + 2)),
            )
    for index, width in enumerate(widths, start=1):
        dimension = sheet.column_dimensions[get_column_letter(index)]
        dimension.width = width
        dimension.hidden = headers[index - 1] in hidden
    sheet.row_dimensions[1].height = 24
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    _append(sheet, headers, header=True, row_number=1)
    status_index = headers.index("Status") if "Status" in headers else None
    for row_number, values in enumerate(rows_factory(), start=2):
        if len(headers) > 10:
            sheet.row_dimensions[row_number].height = 36
        _append(
            sheet,
            values,
            header=False,
            row_number=row_number,
            status_index=status_index,
        )
    last_column = get_column_letter(len(headers))
    sheet.auto_filter.ref = f"A1:{last_column}{max(1, row_count + 1)}"


def _append(
    sheet,
    values: list[object],
    *,
    header: bool,
    row_number: int,
    status_index: int | None = None,
) -> None:
    cells = []
    for index, value in enumerate(values):
        cell = WriteOnlyCell(sheet, value="" if value is None else value)
        if isinstance(value, str):
            cell.data_type = "s"
            cell.number_format = "@"
        cell.font = _HEADER_FONT if header else _BODY_FONT
        cell.alignment = Alignment(
            vertical="top",
            wrap_text=True,
            horizontal="center" if header else "left",
        )
        if header:
            cell.fill = _HEADER_FILL
        elif status_index is not None and index == status_index:
            cell.fill = _STATUS_FILL.get(str(value), _ALT_FILL if row_number % 2 == 0 else PatternFill())
            cell.font = Font(name="Calibri", size=11, bold=True)
        elif row_number % 2 == 0:
            cell.fill = _ALT_FILL
        cells.append(cell)
    sheet.append(cells)


def _input_name(decision: MatchDecision) -> str:
    mapped = decision.mapped_values
    return str(mapped.get("full_name") or "").strip() or " ".join(
        value
        for value in (
            str(mapped.get("first_name") or "").strip(),
            str(mapped.get("last_name") or "").strip(),
        )
        if value
    )


def _advisor_name(advisor: MatchCandidate | None) -> str:
    return (
        " ".join((advisor.first_name, advisor.last_name)).strip()
        if advisor
        else ""
    )


def _location(values: dict[str, str]) -> str:
    city = str(values.get("city") or "").strip()
    state = str(values.get("state") or "").strip()
    zip_code = str(values.get("zip_code") or "").strip()
    locality = ", ".join(value for value in (city, state) if value)
    return " ".join(value for value in (locality, zip_code) if value)


def _advisor_location(advisor: MatchCandidate | None) -> str:
    if not advisor:
        return ""
    return _location(
        {"city": advisor.city, "state": advisor.state, "zip_code": advisor.zip_code}
    )


def _display_width(value: object) -> int:
    text = str(value or "")
    return max((len(line) for line in text.splitlines()), default=0)
