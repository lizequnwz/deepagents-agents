"""Bounded raw-table profiling for agent-directed input interpretation."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from general_agent.advisor_matching.source import sha256_file
from general_agent.config import Settings

_ALIASES = {
    "crd_number": {"crd", "crd number", "finra crd", "advisor crd"},
    "first_name": {"first name", "firstname", "advisor first", "given name"},
    "last_name": {"last name", "lastname", "advisor last", "surname"},
    "full_name": {"name", "full name", "advisor name", "representative"},
    "firm_name": {
        "firm",
        "firm name",
        "company",
        "organization",
        "broker dealer",
    },
    "email": {"email", "email address", "e mail"},
    "city": {"city", "town"},
    "state": {"state", "province"},
    "zip_code": {"zip", "zip code", "postal code", "postcode"},
}


def inspect_advisor_upload(path: Path, settings: Settings) -> dict[str, Any]:
    """Return bounded raw previews and plausible headed/headerless views."""

    suffix = path.suffix.lower()
    if suffix not in {".csv", ".xlsx"}:
        raise ValueError("Advisor matching accepts only CSV or XLSX files.")
    try:
        if suffix == ".csv":
            frames = [(None, _read_csv(path, settings.max_inspect_rows + 1))]
            total_sheets = 1
        else:
            with pd.ExcelFile(path) as excel:
                total_sheets = len(excel.sheet_names)
                frames = [
                    (
                        name,
                        pd.read_excel(
                            excel,
                            sheet_name=name,
                            header=None,
                            dtype=str,
                            keep_default_na=False,
                            nrows=settings.max_inspect_rows + 1,
                        ),
                    )
                    for name in excel.sheet_names[: settings.max_inspect_sheets]
                ]
    except (OSError, UnicodeError, ValueError, BadZipFile, InvalidFileException) as exc:
        raise ValueError(f"The input table could not be read: {exc}") from exc

    sheets = [
        _profile_sheet(name, frame.iloc[:, : settings.max_inspect_columns], settings)
        for name, frame in frames
    ]
    warnings = []
    if total_sheets > settings.max_inspect_sheets:
        warnings.append(
            f"Only the first {settings.max_inspect_sheets} worksheets were profiled."
        )
    return {
        "input_virtual_path": f"/uploads/{path.name}",
        "format": suffix[1:],
        "source_sha256": sha256_file(path),
        "sheets": sheets,
        "warnings": warnings,
    }


def _profile_sheet(
    name: str | None, frame: pd.DataFrame, settings: Settings
) -> dict[str, Any]:
    preview = frame.iloc[: settings.max_inspect_rows]
    rows = [
        (int(index) + 1, [_cell_text(value) for value in row.tolist()])
        for index, row in preview.iterrows()
    ]
    nonblank = [(number, values) for number, values in rows if any(values)]
    ranked = sorted(
        nonblank,
        key=lambda item: (-_header_likelihood(item[1]), item[0]),
    )[:5]
    candidates = [
        _header_candidate(number, values, frame, settings)
        for number, values in ranked
    ]
    return {
        "name": name,
        "preview_row_count": len(rows),
        "preview_truncated": len(frame.index) > settings.max_inspect_rows,
        "preview_rows": [
            {"row_number": number, "values": values}
            for number, values in nonblank[:8]
        ],
        "header_candidates": candidates,
        "headerless": _headerless_view(preview, settings),
    }


def _header_candidate(
    row_number: int,
    headers: list[str],
    frame: pd.DataFrame,
    settings: Settings,
) -> dict[str, Any]:
    samples = frame.iloc[
        row_number : row_number + 3, : settings.max_inspect_columns
    ]
    labels = _unique_labels(headers)
    suggestions: dict[str, list[dict[str, object]]] = {}
    columns = []
    for index, header in enumerate(headers):
        normalized = _normalize_header(header)
        series = samples.iloc[:, index] if index < len(samples.columns) else pd.Series()
        columns.append(
            {
                "index": index,
                "header": header,
                "label": labels[index],
                "non_null_sample": sum(bool(_cell_text(value)) for value in series),
                "pattern": _pattern(series),
            }
        )
        for field, aliases in _ALIASES.items():
            if normalized in aliases:
                suggestions.setdefault(field, []).append(
                    {"index": index, "header": header}
                )
    return {
        "row_number": row_number,
        "columns": columns,
        "sample_rows": [
            {
                "row_number": int(index) + 1,
                "values": {
                    labels[column]: _cell_text(row.iloc[column])
                    for column in range(len(labels))
                },
            }
            for index, row in samples.iterrows()
        ],
        "mapping_suggestions": suggestions,
    }


def _headerless_view(frame: pd.DataFrame, settings: Settings) -> dict[str, Any]:
    labels = [
        f"Column {get_column_letter(index + 1)}"
        for index in range(len(frame.columns))
    ]
    return {
        "columns": [
            {
                "index": index,
                "header": None,
                "label": label,
                "pattern": _pattern(frame.iloc[:, index]),
            }
            for index, label in enumerate(labels)
        ],
        "sample_rows": [
            {
                "row_number": int(index) + 1,
                "values": {
                    labels[column]: _cell_text(row.iloc[column])
                    for column in range(len(labels))
                },
            }
            for index, row in frame.iterrows()
            if any(_cell_text(value) for value in row.tolist())
        ][:3],
    }


def _header_likelihood(values: list[str]) -> float:
    present = [value for value in values if value]
    if not present:
        return 0
    aliases = sum(
        _normalize_header(value) in known
        for value in present
        for known in _ALIASES.values()
    )
    textual = sum(any(character.isalpha() for character in value) for value in present)
    density = len(present) / max(1, len(values))
    uniqueness = len({_normalize_header(value) for value in present}) / len(present)
    return aliases * 10 + textual / len(present) + density + uniqueness


def _pattern(series: pd.Series) -> str:
    values = [_cell_text(value) for value in series if _cell_text(value)][:20]
    if values and sum("@" in value for value in values) >= max(1, len(values) // 2):
        return "email"
    if values and all(value.replace(".0", "").isdigit() for value in values):
        return "numeric"
    return "text"


def _unique_labels(headers: list[str]) -> list[str]:
    used: dict[str, int] = {}
    result = []
    for index, header in enumerate(headers):
        base = header.strip() or f"Column {get_column_letter(index + 1)}"
        used[base] = used.get(base, 0) + 1
        result.append(base if used[base] == 1 else f"{base} [{used[base]}]")
    return result


def _normalize_header(value: str) -> str:
    return " ".join(value.strip().casefold().replace("_", " ").split())


def _cell_text(value: object) -> str:
    return "" if pd.isna(value) else str(value).strip()


def _read_csv(path: Path, nrows: int) -> pd.DataFrame:
    return pd.read_csv(
        path,
        header=None,
        dtype=str,
        keep_default_na=False,
        skip_blank_lines=False,
        nrows=nrows,
        sep=_csv_separator(path),
    )


def _csv_separator(path: Path) -> str:
    with path.open(encoding="utf-8-sig", errors="replace") as handle:
        sample = handle.read(65_536)
    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
    except csv.Error:
        return ","
