"""Validate and load mapped source rows without changing source values."""

from __future__ import annotations

import csv
import hashlib
import json
from dataclasses import dataclass
from pathlib import Path
from zipfile import BadZipFile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils import get_column_letter

from general_agent.advisor_matching import normalization as norm
from general_agent.advisor_matching.schemas import (
    FieldBinding,
    InputMapping,
    InputSummary,
)
from general_agent.advisor_matching.source import sha256_file

_MAPPED_FIELDS = (
    "crd_number",
    "first_name",
    "last_name",
    "full_name",
    "firm_name",
    "email",
    "city",
    "state",
    "zip_code",
)


@dataclass(frozen=True, slots=True)
class LoadedInput:
    rows: list[tuple[int, dict[str, object], dict[str, str]]]
    selected_sheet: str | None
    columns: list[dict[str, object]]
    source_sha256: str
    mapping_fingerprint: str
    summary: InputSummary
    missing_firm_sample: list[dict[str, object]]


def validate_and_load_input(
    path: Path,
    mapping: InputMapping,
    *,
    max_rows: int,
    missing_firm_sample_limit: int = 3,
) -> LoadedInput:
    """Validate the exact table shape, then load bounded nonblank data rows."""

    suffix = path.suffix.lower()
    if suffix not in {".csv", ".xlsx"}:
        raise ValueError("Advisor matching accepts only CSV or XLSX files.")
    selected_sheet = _selected_sheet(path, mapping)
    data_start = mapping.header_row or 0
    read_limit = data_start + max_rows + 1
    raw = _read_raw(
        path,
        sheet_name=selected_sheet,
        nrows=read_limit,
    )
    if mapping.header_row is not None and mapping.header_row > len(raw.index):
        raise ValueError(
            f"Header row {mapping.header_row} does not exist in the selected table."
        )

    observed_headers = (
        [_cell_text(value) for value in raw.iloc[mapping.header_row - 1].tolist()]
        if mapping.header_row is not None
        else [None] * len(raw.columns)
    )
    _validate_references(mapping, observed_headers, len(raw.columns))
    labels = _source_labels(observed_headers, len(raw.columns))
    columns = [
        {
            "index": index,
            "header": observed_headers[index],
            "label": labels[index],
        }
        for index in range(len(raw.columns))
    ]

    data = raw.iloc[data_start:]
    nonblank_flags = [not _row_is_blank(row) for _, row in data.iterrows()]
    last_nonblank = max(
        (index for index, present in enumerate(nonblank_flags) if present),
        default=-1,
    )
    has_later_data = bool(
        len(raw.index) >= read_limit
        and _has_nonblank_after(
            path,
            sheet_name=selected_sheet,
            physical_rows=len(raw.index),
        )
    )
    if last_nonblank >= max_rows or has_later_data:
        raise ValueError(
            f"Input contains more than {max_rows:,} physical data rows; "
            f"limit is {max_rows:,}."
        )

    results: list[tuple[int, dict[str, object], dict[str, str]]] = []
    blank_rows = 0
    missing_firm_sample: list[dict[str, object]] = []
    missing_firm_count = 0
    for raw_index, row in data.iloc[: last_nonblank + 1].iterrows():
        physical_row = int(raw_index) + 1
        if _row_is_blank(row):
            blank_rows += 1
            continue
        source = {
            label: _source_value(row.iloc[index])
            for index, label in enumerate(labels)
        }
        mapped = {
            field: _binding_value(row, getattr(mapping, field))
            if getattr(mapping, field) is not None
            else ""
            for field in _MAPPED_FIELDS
        }
        if _requires_missing_firm_confirmation(mapped):
            missing_firm_count += 1
            if len(missing_firm_sample) < missing_firm_sample_limit:
                missing_firm_sample.append(
                    {
                        "source_row_number": physical_row,
                        "name": _display_name(mapped),
                        "city": mapped.get("city", ""),
                        "state": mapped.get("state", ""),
                    }
                )
        results.append((physical_row, source, mapped))

    if not results:
        raise ValueError("The selected table contains no nonblank data rows.")

    source_sha256 = sha256_file(path)
    mapping_fingerprint = input_mapping_fingerprint(
        source_sha256=source_sha256,
        mapping=mapping,
    )
    summary = InputSummary(
        data_row_count=len(results),
        blank_row_count=blank_rows,
        preamble_row_count=(mapping.header_row - 1)
        if mapping.header_row is not None
        else 0,
        firm_column_missing=mapping.firm_name is None,
        missing_firm_row_count=missing_firm_count,
        missing_firm_confirmation_required=(
            mapping.firm_name is None or missing_firm_count > 0
        ),
    )
    return LoadedInput(
        rows=results,
        selected_sheet=selected_sheet,
        columns=columns,
        source_sha256=source_sha256,
        mapping_fingerprint=mapping_fingerprint,
        summary=summary,
        missing_firm_sample=missing_firm_sample,
    )


def load_input(
    path: Path, mapping: InputMapping, *, max_rows: int
) -> list[tuple[int, dict[str, object], dict[str, str]]]:
    """Compatibility wrapper for developer scripts and deterministic tests."""

    return validate_and_load_input(path, mapping, max_rows=max_rows).rows


def input_mapping_fingerprint(*, source_sha256: str, mapping: InputMapping) -> str:
    canonical = json.dumps(
        mapping.model_dump(mode="json"),
        ensure_ascii=False,
        separators=(",", ":"),
        sort_keys=True,
    )
    return hashlib.sha256(f"{source_sha256}:{canonical}".encode("utf-8")).hexdigest()


def _selected_sheet(path: Path, mapping: InputMapping) -> str | None:
    if path.suffix.lower() == ".csv":
        if mapping.sheet_name is not None:
            raise ValueError("CSV mappings cannot select a worksheet.")
        return None
    try:
        excel = pd.ExcelFile(path)
    except (OSError, ValueError, BadZipFile, InvalidFileException) as exc:
        raise ValueError(f"The input table could not be read: {exc}") from exc
    try:
        if not excel.sheet_names:
            raise ValueError("The Excel workbook contains no worksheets.")
        selected = mapping.sheet_name or excel.sheet_names[0]
        if selected not in excel.sheet_names:
            raise ValueError(f"Worksheet {selected!r} does not exist.")
        return selected
    finally:
        excel.close()


def _read_raw(
    path: Path, *, sheet_name: str | None, nrows: int
) -> pd.DataFrame:
    try:
        if path.suffix.lower() == ".csv":
            return pd.read_csv(
                path,
                header=None,
                dtype=str,
                keep_default_na=False,
                skip_blank_lines=False,
                nrows=nrows,
                sep=_csv_separator(path),
            )
        return pd.read_excel(
            path,
            sheet_name=sheet_name or 0,
            header=None,
            dtype=str,
            keep_default_na=False,
            nrows=nrows,
        )
    except (OSError, UnicodeError, ValueError, BadZipFile, InvalidFileException) as exc:
        raise ValueError(f"The input table could not be read: {exc}") from exc


def _has_nonblank_after(
    path: Path,
    *,
    sheet_name: str | None,
    physical_rows: int,
) -> bool:
    """Detect later records without retaining an unbounded table in memory."""

    if path.suffix.lower() == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as handle:
            reader = csv.reader(handle, delimiter=_csv_separator(path))
            return any(
                any(str(value or "").strip() for value in row)
                for index, row in enumerate(reader)
                if index >= physical_rows
            )
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
        return any(
            any(str(value or "").strip() for value in row)
            for index, row in enumerate(sheet.iter_rows(values_only=True))
            if index >= physical_rows
        )
    finally:
        workbook.close()


def _validate_references(
    mapping: InputMapping,
    observed_headers: list[str | None],
    column_count: int,
) -> None:
    for field, binding in mapping.field_bindings().items():
        for reference in binding.columns:
            if reference.index >= column_count:
                raise ValueError(
                    f"Mapped field {field!r} references missing column index "
                    f"{reference.index}."
                )
            observed = observed_headers[reference.index]
            if observed != reference.header:
                raise ValueError(
                    f"Column mapping no longer matches index {reference.index}: "
                    f"expected {reference.header!r}, observed {observed!r}."
                )


def _binding_value(row: pd.Series, binding: FieldBinding) -> str:
    values = [_cell_text(row.iloc[reference.index]) for reference in binding.columns]
    if binding.combine == "join_space":
        return " ".join(value for value in values if value)
    return next((value for value in values if value), "")


def _source_labels(
    observed_headers: list[str | None], column_count: int
) -> list[str]:
    used: dict[str, int] = {}
    result = []
    for index in range(column_count):
        header = observed_headers[index]
        base = header.strip() if isinstance(header, str) and header.strip() else (
            f"Column {get_column_letter(index + 1)}"
        )
        used[base] = used.get(base, 0) + 1
        result.append(base if used[base] == 1 else f"{base} [{used[base]}]")
    return result


def _requires_missing_firm_confirmation(mapped: dict[str, str]) -> bool:
    return bool(
        _usable_name(mapped)
        and not norm.firm(mapped.get("firm_name"))
        and not norm.crd(mapped.get("crd_number"))
        and not norm.email(mapped.get("email"))
    )


def _usable_name(mapped: dict[str, str]) -> str:
    full = norm.person_name(mapped.get("full_name"))
    if len(full.split()) >= 2:
        return full
    first = norm.first_name(mapped.get("first_name"))
    last = norm.person_name(mapped.get("last_name"))
    return f"{first} {last}".strip() if first and last else ""


def _display_name(mapped: dict[str, str]) -> str:
    return (
        str(mapped.get("full_name") or "").strip()
        or " ".join(
            value
            for value in (
                str(mapped.get("first_name") or "").strip(),
                str(mapped.get("last_name") or "").strip(),
            )
            if value
        )
    )


def _row_is_blank(row: pd.Series) -> bool:
    return all(not _cell_text(value) for value in row.tolist())


def _cell_text(value: object) -> str:
    return "" if pd.isna(value) else str(value).strip()


def _source_value(value: object) -> object:
    return "" if pd.isna(value) else value


def _csv_separator(path: Path) -> str:
    with path.open(encoding="utf-8-sig", errors="replace") as handle:
        sample = handle.read(65_536)
    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
    except csv.Error:
        return ","
