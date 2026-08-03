"""Deterministic immutable input augmentation for user-supplied firm data."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from general_agent.advisor_matching import normalization as norm
from general_agent.advisor_matching.input_loader import (
    _csv_separator,
    validate_and_load_input,
)
from general_agent.advisor_matching.schemas import InputMapping

_FIRM_HEADERS = {
    "firm",
    "firm name",
    "company",
    "company name",
    "organization",
    "broker dealer",
}
_FORMULA_PREFIXES = ("=", "+", "-", "@")


@dataclass(frozen=True, slots=True)
class FirmAugmentation:
    mapping: InputMapping
    firm_name: str
    firm_column_index: int
    firm_column_header: str | None
    rows_updated: int
    selected_sheet: str | None
    source_sha256: str


def augment_input_with_firm(
    source: Path,
    output: Path,
    mapping: InputMapping,
    firm_name: str,
    *,
    max_rows: int,
) -> FirmAugmentation:
    """Create a same-format derived input with one firm value on every data row."""

    firm = _validated_firm_name(firm_name)
    if mapping.firm_name is not None:
        raise ValueError(
            "The interpreted mapping already includes a firm column; use that "
            "source data instead of applying a conversational firm."
        )
    loaded = validate_and_load_input(source, mapping, max_rows=max_rows)
    if _recognized_firm_header_exists(loaded.columns):
        raise ValueError(
            "The selected table already has a recognized firm column; map and "
            "validate that column instead."
        )

    row_numbers = {row_number for row_number, _, _ in loaded.rows}
    if source.suffix.lower() == ".csv":
        column_index = _augment_csv(
            source,
            output,
            header_row=mapping.header_row,
            row_numbers=row_numbers,
            firm_name=firm,
        )
    else:
        column_index = _augment_xlsx(
            source,
            output,
            sheet_name=loaded.selected_sheet,
            header_row=mapping.header_row,
            row_numbers=row_numbers,
            firm_name=firm,
        )
    column_header = "Firm Name" if mapping.header_row is not None else None
    updated_mapping = InputMapping.model_validate(
        {
            **mapping.model_dump(mode="python"),
            "firm_name": {
                "columns": [
                    {"index": column_index, "header": column_header}
                ]
            },
        }
    )
    return FirmAugmentation(
        mapping=updated_mapping,
        firm_name=firm,
        firm_column_index=column_index,
        firm_column_header=column_header,
        rows_updated=len(row_numbers),
        selected_sheet=loaded.selected_sheet,
        source_sha256=loaded.source_sha256,
    )


def _augment_csv(
    source: Path,
    output: Path,
    *,
    header_row: int | None,
    row_numbers: set[int],
    firm_name: str,
) -> int:
    delimiter = _csv_separator(source)
    with source.open(newline="", encoding="utf-8-sig") as handle:
        column_index = max(
            (len(row) for row in csv.reader(handle, delimiter=delimiter)),
            default=0,
        )
    output.parent.mkdir(parents=True, exist_ok=True)
    with (
        source.open(newline="", encoding="utf-8-sig") as source_handle,
        output.open("w", newline="", encoding="utf-8-sig") as output_handle,
    ):
        reader = csv.reader(source_handle, delimiter=delimiter)
        writer = csv.writer(output_handle, delimiter=delimiter)
        for row_number, row in enumerate(reader, start=1):
            row.extend([""] * (column_index + 1 - len(row)))
            if header_row is not None and row_number == header_row:
                row[column_index] = "Firm Name"
            elif row_number in row_numbers:
                row[column_index] = firm_name
            writer.writerow(row)
    return column_index


def _augment_xlsx(
    source: Path,
    output: Path,
    *,
    sheet_name: str | None,
    header_row: int | None,
    row_numbers: set[int],
    firm_name: str,
) -> int:
    workbook = load_workbook(source, data_only=False)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
        column_number = sheet.max_column + 1
        if header_row is not None:
            sheet.cell(header_row, column_number, "Firm Name")
        for row_number in row_numbers:
            sheet.cell(row_number, column_number, firm_name)
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output)
        return column_number - 1
    finally:
        workbook.close()


def _recognized_firm_header_exists(columns: list[dict[str, object]]) -> bool:
    return any(
        _normalize_header(str(column.get("header") or "")) in _FIRM_HEADERS
        for column in columns
    )


def _normalize_header(value: str) -> str:
    return " ".join(value.strip().casefold().replace("_", " ").split())


def _validated_firm_name(value: str) -> str:
    firm = " ".join(str(value or "").split())
    if not firm or not norm.firm(firm):
        raise ValueError("Provide a nonblank, meaningful firm name.")
    if len(firm) > 200:
        raise ValueError("Firm name must be 200 characters or fewer.")
    if firm.startswith(_FORMULA_PREFIXES):
        raise ValueError("Firm name cannot begin with a spreadsheet formula prefix.")
    return firm
