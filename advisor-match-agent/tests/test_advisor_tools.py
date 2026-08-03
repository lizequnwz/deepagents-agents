from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

import general_agent.advisor_workflow as workflow_module
from general_agent.advisor_backend import AdvisorWorkspaceBackend
from general_agent.advisor_matching.source import SyntheticAdvisorReferenceSource
from general_agent.advisor_matching.workbook import verify_match_workbook
from general_agent.advisor_tools import build_advisor_tools
from general_agent.observability import configure_logging, shutdown_logging
from general_agent.store import Store
from general_agent.workspace import Workspace


def _source():
    return SyntheticAdvisorReferenceSource(
        Path(__file__).parents[1]
        / "general_agent/advisor_matching/data/master_advisors.csv"
    )


class CountingAdvisorSource:
    source_kind = "synthetic"
    schema_version = "1"

    def __init__(self) -> None:
        self.calls = 0

    def iter_records(self):
        self.calls += 1
        yield from _source().iter_records()


def _tools(settings, workspace, backend, store, *, source=None):
    return {
        item.name: item
        for item in build_advisor_tools(
            settings=settings,
            workspace=workspace,
            backend=backend,
            store=store,
            advisor_source=source or _source(),
        )
    }


@pytest.fixture
def operational_log(settings):
    configure_logging(settings)
    try:
        yield settings.api_log
    finally:
        shutdown_logging()


def _upload(
    workspace: Workspace,
    store: Store,
    *,
    corp_id: str,
    conversation_id: str,
    run_id: str,
    name: str,
    content: bytes,
) -> str:
    attachment, protected = workspace.upload(
        corp_id=corp_id,
        conversation_id=conversation_id,
        original_name=name,
        content_type="text/csv",
        source=BytesIO(content),
        max_bytes=1024 * 1024,
    )
    store.add_attachment(run_id, attachment, protected, corp_id=corp_id)
    return attachment.attachment_id


@pytest.mark.asyncio
async def test_three_stage_match_review_and_regeneration(
    settings, operational_log
) -> None:
    corp_id = "A123456"
    workspace = Workspace(settings.data_root)
    store = Store(settings.application_db, settings.data_root)
    conversation_id = store.create_conversation(corp_id=corp_id).conversation_id
    first_run_id, _ = store.create_run(
        conversation_id, "match", corp_id=corp_id
    )
    attachment_id = _upload(
        workspace,
        store,
        corp_id=corp_id,
        conversation_id=conversation_id,
        run_id=first_run_id,
        name="advisors.csv",
        content=(
            b"Advisor Name,Firm,City,State\n"
            b"John Smith,Northstar Wealth Partners,Boston,MA\n"
        ),
    )
    backend = AdvisorWorkspaceBackend(settings.runtime_root)
    advisor_source = CountingAdvisorSource()
    tools = _tools(
        settings, workspace, backend, store, source=advisor_source
    )
    mapping = {
        "full_name": {"columns": [{"index": 0, "header": "Advisor Name"}]},
        "firm_name": {"columns": [{"index": 1, "header": "Firm"}]},
        "city": {"columns": [{"index": 2, "header": "City"}]},
        "state": {"columns": [{"index": 3, "header": "State"}]},
    }

    async with backend.run_scope(first_run_id, corp_id, conversation_id):
        profile = await tools["inspect_advisor_upload"].ainvoke(
            {"attachment_id": attachment_id}
        )
        candidate = profile["sheets"][0]["header_candidates"][0]
        assert candidate["row_number"] == 1
        assert candidate["columns"][0]["header"] == "Advisor Name"

        validation = await tools["validate_advisor_input"].ainvoke(
            {
                "attachment_id": attachment_id,
                "mapping": mapping,
            }
        )
        assert validation["input_summary"]["data_row_count"] == 1
        assert validation["input_summary"]["missing_firm_confirmation_required"] is False

        result = await tools["create_advisor_match"].ainvoke(
            {
                "attachment_id": attachment_id,
                "mapping": mapping,
                "mapping_fingerprint": validation["mapping_fingerprint"],
            }
        )
        manifest = result["reference"]
        assert manifest["row_count"] == 40
        assert manifest["reference_snapshot_id"].startswith("ars_")
        assert "snapshot_path" not in manifest
        assert "STREET_ADDRESS" not in manifest["columns"]
        stored_reference = store.get_advisor_reference_snapshot(
            manifest["reference_snapshot_id"],
            corp_id=corp_id,
            conversation_id=conversation_id,
        )
        assert Path(stored_reference["snapshot_path"]).is_file()
        with pytest.raises(KeyError):
            store.get_advisor_reference_snapshot(
                manifest["reference_snapshot_id"], corp_id="B654321"
            )

        assert result["counts"]["ambiguous_match"] == 1
        assert result["interpreted_mapping"] == validation["mapping"]
        current = await tools["get_current_advisor_match"].ainvoke({})
        assert current["match_session_id"] == result["match_session_id"]
        assert current["counts"] == result["counts"]
        assert current["interpreted_mapping"] == result["interpreted_mapping"]

        repeated = await tools["create_advisor_match"].ainvoke(
            {
                "attachment_id": attachment_id,
                "mapping": mapping,
                "mapping_fingerprint": validation["mapping_fingerprint"],
            }
        )
        assert repeated["reference"]["reference_snapshot_id"] == manifest[
            "reference_snapshot_id"
        ]
        assert advisor_source.calls == 1

        page = await tools["list_advisor_match_results"].ainvoke(
            {
                "match_session_id": result["match_session_id"],
                "status": "Ambiguous Match",
            }
        )
        item = page["items"][0]
        assert "internal_score" not in item["candidates"][0]
        assert item["candidates"][0]["supporting_evidence"]
        selected_crd = item["candidates"][0]["crd_number"]
        updated = await tools["apply_advisor_match_decisions"].ainvoke(
            {
                "match_session_id": result["match_session_id"],
                "decisions": [
                    {
                        "review_item_id": item["review_item_id"],
                        "action": "confirm_candidate",
                        "crd_number": f"  {selected_crd}  ",
                    }
                ],
                "approve_session": True,
            }
        )
        assert updated["counts"] == {
            "matched": 1,
            "ambiguous_match": 0,
            "no_match": 0,
        }

    session = store.get_advisor_match_session(
        result["match_session_id"], corp_id=corp_id
    )
    assert session["status"] == "Approved"
    assert session["decisions"][0]["automated_status"] == "Ambiguous Match"
    audit = store._connection.execute(
        "SELECT prior_decision_json, new_decision_json "
        "FROM advisor_match_review_decisions WHERE session_id=?",
        (result["match_session_id"],),
    ).fetchone()
    assert audit is not None
    with pytest.raises(KeyError):
        store.get_advisor_match_session(
            result["match_session_id"], corp_id="B654321"
        )
    output, _ = store.artifact_path(updated["output_artifact_id"], corp_id=corp_id)
    assert verify_match_workbook(output, expected_rows=1)["matched"] == 1
    artifacts = store.get_conversation(conversation_id, corp_id=corp_id).turns[0].artifacts
    assert [artifact.revision for artifact in artifacts] == [1, 1, 2]
    store.close()
    log_contents = operational_log.read_text(encoding="utf-8")
    assert "agent.artifact.build_started" in log_contents
    assert "agent.artifact.published" in log_contents
    assert f"match_session_id={result['match_session_id']}" in log_contents
    assert "John Smith" not in log_contents


@pytest.mark.asyncio
async def test_create_match_accepts_opaque_master_and_input_crd(settings) -> None:
    corp_id = "A123456"
    workspace = Workspace(settings.data_root)
    store = Store(settings.application_db, settings.data_root)
    conversation_id = store.create_conversation(corp_id=corp_id).conversation_id
    run_id, _ = store.create_run(conversation_id, "match", corp_id=corp_id)
    attachment_id = _upload(
        workspace,
        store,
        corp_id=corp_id,
        conversation_id=conversation_id,
        run_id=run_id,
        name="opaque-crd.csv",
        content=b"CRD,Firm\n FSA_ID:111 ,Example Advisory\n",
    )
    master = settings.data_root / "opaque-master.csv"
    master.write_text(
        "CRD_NUMBER,FIRST_NAME,LAST_NAME,FIRM_NAME,EMAIL,CITY,STATE,ZIP_CODE\n"
        " FSA_ID:111 ,Jane,Doe,Example Advisory,,,,\n",
        encoding="utf-8",
    )
    backend = AdvisorWorkspaceBackend(settings.runtime_root)
    tools = _tools(
        settings,
        workspace,
        backend,
        store,
        source=SyntheticAdvisorReferenceSource(master),
    )
    mapping = {
        "crd_number": {"columns": [{"index": 0, "header": "CRD"}]},
        "firm_name": {"columns": [{"index": 1, "header": "Firm"}]},
    }

    async with backend.run_scope(run_id, corp_id, conversation_id):
        validation = await tools["validate_advisor_input"].ainvoke(
            {"attachment_id": attachment_id, "mapping": mapping}
        )
        result = await tools["create_advisor_match"].ainvoke(
            {
                "attachment_id": attachment_id,
                "mapping": mapping,
                "mapping_fingerprint": validation["mapping_fingerprint"],
            }
        )
        page = await tools["list_advisor_match_results"].ainvoke(
            {"match_session_id": result["match_session_id"], "status": "matched"}
        )

    assert result["counts"] == {
        "matched": 1,
        "ambiguous_match": 0,
        "no_match": 0,
    }
    assert page["items"][0]["matched_advisor"]["crd_number"] == "FSA_ID:111"
    store.close()


@pytest.mark.asyncio
async def test_duplicate_master_crds_return_controlled_blocker(settings) -> None:
    corp_id = "A123456"
    workspace = Workspace(settings.data_root)
    store = Store(settings.application_db, settings.data_root)
    conversation_id = store.create_conversation(corp_id=corp_id).conversation_id
    run_id, _ = store.create_run(conversation_id, "match", corp_id=corp_id)
    attachment_id = _upload(
        workspace,
        store,
        corp_id=corp_id,
        conversation_id=conversation_id,
        run_id=run_id,
        name="input.csv",
        content=b"CRD,Firm\nDUP-1,Example Advisory\n",
    )
    master = settings.data_root / "duplicate-master.csv"
    master.write_text(
        "CRD_NUMBER,FIRST_NAME,LAST_NAME,FIRM_NAME,EMAIL,CITY,STATE,ZIP_CODE\n"
        " DUP-1 ,Jane,Doe,Example Advisory,,,,\n"
        "DUP-1,Janet,Doe,Example Advisory,,,,\n"
        "DUP-1,Jamie,Doe,Example Advisory,,,,\n"
        "DUP-2,Alex,Smith,Example Advisory,,,,\n"
        " DUP-2 ,Avery,Smith,Example Advisory,,,,\n",
        encoding="utf-8",
    )
    backend = AdvisorWorkspaceBackend(settings.runtime_root)
    tools = _tools(
        settings,
        workspace,
        backend,
        store,
        source=SyntheticAdvisorReferenceSource(master),
    )
    mapping = {
        "crd_number": {"columns": [{"index": 0, "header": "CRD"}]},
        "firm_name": {"columns": [{"index": 1, "header": "Firm"}]},
    }
    async with backend.run_scope(run_id, corp_id, conversation_id):
        validation = await tools["validate_advisor_input"].ainvoke(
            {"attachment_id": attachment_id, "mapping": mapping}
        )
        result = await tools["create_advisor_match"].ainvoke(
            {
                "attachment_id": attachment_id,
                "mapping": mapping,
                "mapping_fingerprint": validation["mapping_fingerprint"],
            }
        )

    assert result["workflow_status"] == "blocked"
    assert result["blocker_code"] == "DUPLICATE_REFERENCE_CRD"
    assert result["duplicate_crd_count"] == 2
    assert result["duplicate_crds"] == [
        {"crd_number": "DUP-1", "occurrences": 3},
        {"crd_number": "DUP-2", "occurrences": 2},
    ]
    assert "correct" in result["next_action"]
    assert "upload" not in result["message"].casefold()
    with pytest.raises(KeyError):
        store.get_advisor_reference_snapshot_for_attachment(
            attachment_id, corp_id=corp_id, conversation_id=conversation_id
        )
    with pytest.raises(KeyError):
        store.get_latest_advisor_match_session(conversation_id, corp_id=corp_id)
    assert not any(
        workspace.category_root(corp_id, "advisor_references").iterdir()
    )
    assert not any(workspace.category_root(corp_id, "artifacts").iterdir())
    store.close()


@pytest.mark.asyncio
async def test_invalid_master_data_returns_controlled_source_blocker(settings) -> None:
    corp_id = "A123456"
    workspace = Workspace(settings.data_root)
    store = Store(settings.application_db, settings.data_root)
    conversation_id = store.create_conversation(corp_id=corp_id).conversation_id
    run_id, _ = store.create_run(conversation_id, "match", corp_id=corp_id)
    attachment_id = _upload(
        workspace,
        store,
        corp_id=corp_id,
        conversation_id=conversation_id,
        run_id=run_id,
        name="input.csv",
        content=b"CRD\nDUP-1\n",
    )
    master = settings.data_root / "invalid-master.csv"
    master.write_text(
        "CRD_NUMBER,FIRST_NAME,LAST_NAME,FIRM_NAME,EMAIL,CITY,STATE,ZIP_CODE\n"
        ",Jane,Doe,Example Advisory,,,,\n",
        encoding="utf-8",
    )
    backend = AdvisorWorkspaceBackend(settings.runtime_root)
    tools = _tools(
        settings,
        workspace,
        backend,
        store,
        source=SyntheticAdvisorReferenceSource(master),
    )
    mapping = {"crd_number": {"columns": [{"index": 0, "header": "CRD"}]}}
    async with backend.run_scope(run_id, corp_id, conversation_id):
        validation = await tools["validate_advisor_input"].ainvoke(
            {"attachment_id": attachment_id, "mapping": mapping}
        )
        result = await tools["create_advisor_match"].ainvoke(
            {
                "attachment_id": attachment_id,
                "mapping": mapping,
                "mapping_fingerprint": validation["mapping_fingerprint"],
            }
        )

    assert result == {
        "workflow_status": "blocked",
        "blocker_code": "REFERENCE_DATA_INVALID",
        "message": (
            "The authoritative advisor source failed validation and must be "
            "corrected before matching."
        ),
        "duplicate_crd_count": 0,
        "duplicate_crds": [],
        "next_action": "correct_authoritative_source",
    }
    assert "upload" not in result["message"].casefold()
    with pytest.raises(KeyError):
        store.get_latest_advisor_match_session(conversation_id, corp_id=corp_id)
    store.close()


@pytest.mark.asyncio
async def test_mapping_correction_reuses_completed_attachment_snapshot(settings) -> None:
    corp_id = "A123456"
    workspace = Workspace(settings.data_root)
    store = Store(settings.application_db, settings.data_root)
    conversation_id = store.create_conversation(corp_id=corp_id).conversation_id
    run_id, _ = store.create_run(conversation_id, "match", corp_id=corp_id)
    attachment_id = _upload(
        workspace,
        store,
        corp_id=corp_id,
        conversation_id=conversation_id,
        run_id=run_id,
        name="alternate-name-columns.csv",
        content=(
            b"Name,First,Last,Firm\n"
            b"Avery Stone,Avery,Stone,Northstar Wealth Partners\n"
        ),
    )
    backend = AdvisorWorkspaceBackend(settings.runtime_root)
    advisor_source = CountingAdvisorSource()
    tools = _tools(settings, workspace, backend, store, source=advisor_source)
    full_name_mapping = {
        "full_name": {"columns": [{"index": 0, "header": "Name"}]},
        "firm_name": {"columns": [{"index": 3, "header": "Firm"}]},
    }
    split_name_mapping = {
        "first_name": {"columns": [{"index": 1, "header": "First"}]},
        "last_name": {"columns": [{"index": 2, "header": "Last"}]},
        "firm_name": {"columns": [{"index": 3, "header": "Firm"}]},
    }

    async with backend.run_scope(run_id, corp_id, conversation_id):
        first_validation = await tools["validate_advisor_input"].ainvoke(
            {"attachment_id": attachment_id, "mapping": full_name_mapping}
        )
        first = await tools["create_advisor_match"].ainvoke(
            {
                "attachment_id": attachment_id,
                "mapping": full_name_mapping,
                "mapping_fingerprint": first_validation["mapping_fingerprint"],
            }
        )
        corrected_validation = await tools["validate_advisor_input"].ainvoke(
            {"attachment_id": attachment_id, "mapping": split_name_mapping}
        )
        corrected = await tools["create_advisor_match"].ainvoke(
            {
                "attachment_id": attachment_id,
                "mapping": split_name_mapping,
                "mapping_fingerprint": corrected_validation["mapping_fingerprint"],
            }
        )

    assert corrected["reference"]["reference_snapshot_id"] == first["reference"][
        "reference_snapshot_id"
    ]
    assert advisor_source.calls == 1
    assert first["counts"] == corrected["counts"] == {
        "matched": 1,
        "ambiguous_match": 0,
        "no_match": 0,
    }
    store.close()


@pytest.mark.asyncio
async def test_match_retry_after_failure_reuses_completed_snapshot(
    settings, monkeypatch
) -> None:
    corp_id = "A123456"
    workspace = Workspace(settings.data_root)
    store = Store(settings.application_db, settings.data_root)
    conversation_id = store.create_conversation(corp_id=corp_id).conversation_id
    run_id, _ = store.create_run(conversation_id, "match", corp_id=corp_id)
    attachment_id = _upload(
        workspace,
        store,
        corp_id=corp_id,
        conversation_id=conversation_id,
        run_id=run_id,
        name="retry.csv",
        content=b"Name,Firm\nAvery Stone,Northstar Wealth Partners\n",
    )
    backend = AdvisorWorkspaceBackend(settings.runtime_root)
    advisor_source = CountingAdvisorSource()
    tools = _tools(settings, workspace, backend, store, source=advisor_source)
    mapping = {
        "full_name": {"columns": [{"index": 0, "header": "Name"}]},
        "firm_name": {"columns": [{"index": 1, "header": "Firm"}]},
    }
    real_matcher = workflow_module.run_matching
    attempts = 0

    def fail_once(rows, index):
        nonlocal attempts
        attempts += 1
        if attempts == 1:
            raise RuntimeError("simulated post-snapshot match failure")
        return real_matcher(rows, index)

    monkeypatch.setattr(workflow_module, "run_matching", fail_once)
    async with backend.run_scope(run_id, corp_id, conversation_id):
        validation = await tools["validate_advisor_input"].ainvoke(
            {"attachment_id": attachment_id, "mapping": mapping}
        )
        request = {
            "attachment_id": attachment_id,
            "mapping": mapping,
            "mapping_fingerprint": validation["mapping_fingerprint"],
        }
        with pytest.raises(RuntimeError, match="post-snapshot"):
            await tools["create_advisor_match"].ainvoke(request)
        result = await tools["create_advisor_match"].ainvoke(request)

    assert result["counts"]["matched"] == 1
    assert advisor_source.calls == 1
    assert attempts == 2
    store.close()


@pytest.mark.asyncio
async def test_missing_firm_requires_explicit_continue_and_fingerprint(settings) -> None:
    corp_id = "A123456"
    workspace = Workspace(settings.data_root)
    store = Store(settings.application_db, settings.data_root)
    conversation_id = store.create_conversation(corp_id=corp_id).conversation_id
    first_run_id, _ = store.create_run(
        conversation_id, "match", corp_id=corp_id
    )
    attachment_id = _upload(
        workspace,
        store,
        corp_id=corp_id,
        conversation_id=conversation_id,
        run_id=first_run_id,
        name="name-only.csv",
        content=b"Name\nRobert Mercer\n",
    )
    backend = AdvisorWorkspaceBackend(settings.runtime_root)
    tools = _tools(settings, workspace, backend, store)
    mapping = {"full_name": {"columns": [{"index": 0, "header": "Name"}]}}
    async with backend.run_scope(first_run_id, corp_id, conversation_id):
        validation = await tools["validate_advisor_input"].ainvoke(
            {"attachment_id": attachment_id, "mapping": mapping}
        )
        assert validation["input_summary"]["missing_firm_row_count"] == 1
        clarification = await tools["create_advisor_match"].ainvoke(
            {
                "attachment_id": attachment_id,
                "mapping": mapping,
                "mapping_fingerprint": validation["mapping_fingerprint"],
            }
        )
        assert clarification["workflow_status"] == "firm_clarification_required"
        assert clarification["reason"] == "missing_firm"
        assert clarification["allowed_resolutions"] == [
            "override_all",
            "continue_without_firm",
        ]

    store.finish_run(first_run_id, "completed", corp_id=corp_id)
    second_run_id, _ = store.create_run(
        conversation_id, "No firm is available; continue", corp_id=corp_id
    )
    async with backend.run_scope(second_run_id, corp_id, conversation_id):
        checkpoint = await tools["get_current_advisor_input"].ainvoke({})
        assert checkpoint["mapping_fingerprint"] == validation["mapping_fingerprint"]
        base = {
            "attachment_id": attachment_id,
            "mapping": mapping,
        }
        with pytest.raises(ValueError, match="validate it again"):
            await tools["create_advisor_match"].ainvoke(
                {**base, "mapping_fingerprint": "0" * 64}
            )
        result = await tools["create_advisor_match"].ainvoke(
            {
                **base,
                "mapping_fingerprint": validation["mapping_fingerprint"],
                "firm_resolution": "continue_without_firm",
            }
        )
        assert result["workflow_status"] == "match_created"
        assert result["counts"]["ambiguous_match"] == 1
    store.close()


@pytest.mark.asyncio
async def test_missing_firm_with_crd_proceeds_without_clarification(settings) -> None:
    corp_id = "A123456"
    workspace = Workspace(settings.data_root)
    store = Store(settings.application_db, settings.data_root)
    conversation_id = store.create_conversation(corp_id=corp_id).conversation_id
    run_id, _ = store.create_run(conversation_id, "match", corp_id=corp_id)
    attachment_id = _upload(
        workspace,
        store,
        corp_id=corp_id,
        conversation_id=conversation_id,
        run_id=run_id,
        name="crd-only.csv",
        content=b"CRD\n99000001\n",
    )
    backend = AdvisorWorkspaceBackend(settings.runtime_root)
    tools = _tools(settings, workspace, backend, store)
    mapping = {"crd_number": {"columns": [{"index": 0, "header": "CRD"}]}}
    async with backend.run_scope(run_id, corp_id, conversation_id):
        validation = await tools["validate_advisor_input"].ainvoke(
            {"attachment_id": attachment_id, "mapping": mapping}
        )
        result = await tools["create_advisor_match"].ainvoke(
            {
                "attachment_id": attachment_id,
                "mapping": mapping,
                "mapping_fingerprint": validation["mapping_fingerprint"],
            }
        )
    assert validation["input_summary"]["firm_column_missing"] is True
    assert validation["input_summary"]["missing_firm_confirmation_required"] is False
    assert result["workflow_status"] == "match_created"
    assert result["counts"]["matched"] == 1
    store.close()


@pytest.mark.asyncio
async def test_same_turn_firm_override_matches_without_derived_input(
    settings,
) -> None:
    corp_id = "A123456"
    workspace = Workspace(settings.data_root)
    store = Store(settings.application_db, settings.data_root)
    conversation_id = store.create_conversation(corp_id=corp_id).conversation_id
    first_run_id, _ = store.create_run(
        conversation_id,
        "Match these advisors; all advisors are with Cedar Grove Advisory",
        corp_id=corp_id,
    )
    original = (
        b"Name,City,State\n"
        b"Robert Mercer,Richmond,VA\n"
        b"Unknown Person,Richmond,VA\n"
    )
    attachment_id = _upload(
        workspace,
        store,
        corp_id=corp_id,
        conversation_id=conversation_id,
        run_id=first_run_id,
        name="name-and-location.csv",
        content=original,
    )
    backend = AdvisorWorkspaceBackend(settings.runtime_root)
    tools = _tools(settings, workspace, backend, store)
    mapping = {
        "full_name": {"columns": [{"index": 0, "header": "Name"}]},
        "city": {"columns": [{"index": 1, "header": "City"}]},
        "state": {"columns": [{"index": 2, "header": "State"}]},
    }

    async with backend.run_scope(first_run_id, corp_id, conversation_id):
        validation = await tools["validate_advisor_input"].ainvoke(
            {"attachment_id": attachment_id, "mapping": mapping}
        )
        assert validation["input_summary"]["missing_firm_confirmation_required"] is True
        result = await tools["create_advisor_match"].ainvoke(
            {
                "attachment_id": attachment_id,
                "mapping": mapping,
                "mapping_fingerprint": validation["mapping_fingerprint"],
                "all_rows_firm": "Cedar Grove Advisory",
            }
        )
        assert result["workflow_status"] == "match_created"
        assert result["counts"] == {
            "matched": 1,
            "ambiguous_match": 0,
            "no_match": 1,
        }
        assert result["source_transformation"]["rows_updated"] == 2

    session = store.get_advisor_match_session(result["match_session_id"], corp_id=corp_id)
    assert session["source_attachment_id"] == attachment_id
    assert session["source_transformation"]["source_attachment_id"] == attachment_id
    assert session["source_transformation"]["type"] == "session_firm_override"
    original_path, _, original_sha256 = store.attachment_path(
        attachment_id, corp_id=corp_id, conversation_id=conversation_id
    )
    assert original_path.read_bytes() == original
    assert original_sha256 == validation["source_sha256"]
    workbook_path, _ = store.artifact_path(result["output_artifact_id"], corp_id=corp_id)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        summary = dict(workbook["Run Summary"].iter_rows(min_row=2, values_only=True))
        assert summary["Input Transformation"] == "User-confirmed session firm override"
        assert summary["User-Supplied Firm"] == "Cedar Grove Advisory"
        assert summary["Rows With Firm Override"] == 2
        matched_headers = [cell.value for cell in workbook["Matched"][1]]
        input_firm_index = matched_headers.index("Input Firm") + 1
        assert workbook["Matched"].cell(2, input_firm_index).value == (
            "Cedar Grove Advisory"
        )
        review_headers = [cell.value for cell in workbook["Review Required"][1]]
        review_firm_index = review_headers.index("Input Firm") + 1
        assert workbook["Review Required"].cell(2, review_firm_index).value == (
            "Cedar Grove Advisory"
        )
        original_headers = [cell.value for cell in workbook["Original Input"][1]]
        assert "Firm" not in original_headers
        assert "Firm Name" not in original_headers
    finally:
        workbook.close()
    attachments = store.get_conversation(
        conversation_id, corp_id=corp_id
    ).turns[0].attachments
    assert [item.attachment_id for item in attachments] == [attachment_id]
    store.close()


@pytest.mark.asyncio
async def test_existing_firm_discrepancy_requires_explicit_resolution(settings) -> None:
    corp_id = "A123456"
    workspace = Workspace(settings.data_root)
    store = Store(settings.application_db, settings.data_root)
    conversation_id = store.create_conversation(corp_id=corp_id).conversation_id
    first_run_id, _ = store.create_run(
        conversation_id,
        "All advisors are with Cedar Grove Advisory",
        corp_id=corp_id,
    )
    attachment_id = _upload(
        workspace,
        store,
        corp_id=corp_id,
        conversation_id=conversation_id,
        run_id=first_run_id,
        name="firm-conflict.csv",
        content=b"Name,Firm\nRobert Mercer,Different Advisory\n",
    )
    backend = AdvisorWorkspaceBackend(settings.runtime_root)
    advisor_source = CountingAdvisorSource()
    tools = _tools(settings, workspace, backend, store, source=advisor_source)
    mapping = {
        "full_name": {"columns": [{"index": 0, "header": "Name"}]},
        "firm_name": {"columns": [{"index": 1, "header": "Firm"}]},
    }
    async with backend.run_scope(first_run_id, corp_id, conversation_id):
        validation = await tools["validate_advisor_input"].ainvoke(
            {"attachment_id": attachment_id, "mapping": mapping}
        )
        clarification = await tools["create_advisor_match"].ainvoke(
            {
                "attachment_id": attachment_id,
                "mapping": mapping,
                "mapping_fingerprint": validation["mapping_fingerprint"],
                "all_rows_firm": "Cedar Grove Advisory",
            }
        )
    assert clarification["workflow_status"] == "firm_clarification_required"
    assert clarification["reason"] == "firm_conflict"
    assert clarification["source_firm_sample"] == ["Different Advisory"]
    assert clarification["allowed_resolutions"] == ["use_source", "override_all"]
    assert advisor_source.calls == 0
    with pytest.raises(KeyError):
        store.get_advisor_reference_snapshot_for_attachment(
            attachment_id, corp_id=corp_id, conversation_id=conversation_id
        )

    store.finish_run(first_run_id, "completed", corp_id=corp_id)
    source_run_id, _ = store.create_run(
        conversation_id, "Use the source firm values", corp_id=corp_id
    )
    async with backend.run_scope(source_run_id, corp_id, conversation_id):
        source_result = await tools["create_advisor_match"].ainvoke(
            {
                "attachment_id": attachment_id,
                "mapping": mapping,
                "mapping_fingerprint": validation["mapping_fingerprint"],
                "firm_resolution": "use_source",
            }
        )
    assert source_result["workflow_status"] == "match_created"
    assert source_result["source_transformation"] == {}

    store.finish_run(source_run_id, "completed", corp_id=corp_id)
    incomplete_run_id, _ = store.create_run(
        conversation_id, "Override all rows", corp_id=corp_id
    )
    async with backend.run_scope(incomplete_run_id, corp_id, conversation_id):
        with pytest.raises(ValueError, match="current user message"):
            await tools["create_advisor_match"].ainvoke(
                {
                    "attachment_id": attachment_id,
                    "mapping": mapping,
                    "mapping_fingerprint": validation["mapping_fingerprint"],
                    "all_rows_firm": "Cedar Grove Advisory",
                    "firm_resolution": "override_all",
                }
            )

    store.finish_run(incomplete_run_id, "completed", corp_id=corp_id)
    override_run_id, _ = store.create_run(
        conversation_id,
        "Override all rows with Cedar Grove Advisory",
        corp_id=corp_id,
    )
    async with backend.run_scope(override_run_id, corp_id, conversation_id):
        override_result = await tools["create_advisor_match"].ainvoke(
            {
                "attachment_id": attachment_id,
                "mapping": mapping,
                "mapping_fingerprint": validation["mapping_fingerprint"],
                "all_rows_firm": "Cedar Grove Advisory",
                "firm_resolution": "override_all",
            }
        )
    assert override_result["workflow_status"] == "match_created"
    assert override_result["source_transformation"]["type"] == (
        "session_firm_override"
    )
    assert advisor_source.calls == 1
    store.close()


@pytest.mark.parametrize(
    ("content", "reason"),
    [
        (b"Name,Firm\nRobert Mercer,\n", "blank_source_firms"),
        (
            b"Name,Firm\nRobert Mercer,Cedar Grove Advisory\n"
            b"Avery Stone,Northstar Wealth Partners\n",
            "mixed_source_firms",
        ),
    ],
)
@pytest.mark.asyncio
async def test_blank_and_mixed_source_firms_return_bounded_clarification(
    settings, content, reason
) -> None:
    corp_id = "A123456"
    workspace = Workspace(settings.data_root)
    store = Store(settings.application_db, settings.data_root)
    conversation_id = store.create_conversation(corp_id=corp_id).conversation_id
    run_id, _ = store.create_run(
        conversation_id,
        "All advisors are with Cedar Grove Advisory",
        corp_id=corp_id,
    )
    attachment_id = _upload(
        workspace,
        store,
        corp_id=corp_id,
        conversation_id=conversation_id,
        run_id=run_id,
        name="firm-discrepancy.csv",
        content=content,
    )
    backend = AdvisorWorkspaceBackend(settings.runtime_root)
    source = CountingAdvisorSource()
    tools = _tools(settings, workspace, backend, store, source=source)
    mapping = {
        "full_name": {"columns": [{"index": 0, "header": "Name"}]},
        "firm_name": {"columns": [{"index": 1, "header": "Firm"}]},
    }
    async with backend.run_scope(run_id, corp_id, conversation_id):
        validation = await tools["validate_advisor_input"].ainvoke(
            {"attachment_id": attachment_id, "mapping": mapping}
        )
        result = await tools["create_advisor_match"].ainvoke(
            {
                "attachment_id": attachment_id,
                "mapping": mapping,
                "mapping_fingerprint": validation["mapping_fingerprint"],
                "all_rows_firm": "Cedar Grove Advisory",
            }
        )
    assert result["workflow_status"] == "firm_clarification_required"
    assert result["reason"] == reason
    assert len(result["affected_row_sample"]) <= 5
    assert source.calls == 0
    store.close()


@pytest.mark.asyncio
async def test_normalized_existing_firm_agreement_proceeds_without_override(
    settings,
) -> None:
    corp_id = "A123456"
    workspace = Workspace(settings.data_root)
    store = Store(settings.application_db, settings.data_root)
    conversation_id = store.create_conversation(corp_id=corp_id).conversation_id
    run_id, _ = store.create_run(
        conversation_id,
        "All advisors are with Cedar Grove Advisory",
        corp_id=corp_id,
    )
    attachment_id = _upload(
        workspace,
        store,
        corp_id=corp_id,
        conversation_id=conversation_id,
        run_id=run_id,
        name="firm-agreement.csv",
        content=b"Name,Firm\nRobert Mercer,Cedar Grove Advisory LLC\n",
    )
    backend = AdvisorWorkspaceBackend(settings.runtime_root)
    tools = _tools(settings, workspace, backend, store)
    mapping = {
        "full_name": {"columns": [{"index": 0, "header": "Name"}]},
        "firm_name": {"columns": [{"index": 1, "header": "Firm"}]},
    }
    async with backend.run_scope(run_id, corp_id, conversation_id):
        validation = await tools["validate_advisor_input"].ainvoke(
            {"attachment_id": attachment_id, "mapping": mapping}
        )
        result = await tools["create_advisor_match"].ainvoke(
            {
                "attachment_id": attachment_id,
                "mapping": mapping,
                "mapping_fingerprint": validation["mapping_fingerprint"],
                "all_rows_firm": "Cedar Grove Advisory",
            }
        )
    assert result["workflow_status"] == "match_created"
    assert result["source_transformation"] == {}
    store.close()


@pytest.mark.asyncio
async def test_legacy_derived_attachment_can_match_and_regenerate(settings) -> None:
    corp_id = "A123456"
    workspace = Workspace(settings.data_root)
    store = Store(settings.application_db, settings.data_root)
    conversation_id = store.create_conversation(corp_id=corp_id).conversation_id
    run_id, _ = store.create_run(conversation_id, "match", corp_id=corp_id)
    original_id = _upload(
        workspace,
        store,
        corp_id=corp_id,
        conversation_id=conversation_id,
        run_id=run_id,
        name="legacy.csv",
        content=b"Name,City,State\nJohn Smith,Boston,MA\n",
    )
    _, _, original_sha256 = store.attachment_path(original_id, corp_id=corp_id)
    derived_content = (
        b"Name,City,State,Firm Name\n"
        b"John Smith,Boston,MA,Northstar Wealth Partners\n"
    )
    derived, protected = workspace.upload(
        corp_id=corp_id,
        conversation_id=conversation_id,
        original_name="legacy_with_firm.csv",
        content_type="text/csv",
        source=BytesIO(derived_content),
        max_bytes=1024 * 1024,
    )
    store.add_attachment(
        run_id,
        derived,
        protected,
        corp_id=corp_id,
        derived_from_attachment_id=original_id,
        transformation={
            "type": "bulk_firm_augmentation",
            "source_attachment_id": original_id,
            "source_sha256": original_sha256,
            "firm_name": "Northstar Wealth Partners",
            "rows_updated": 1,
            "selected_sheet": None,
            "firm_column_index": 3,
            "firm_column_header": "Firm Name",
        },
    )
    backend = AdvisorWorkspaceBackend(settings.runtime_root)
    tools = _tools(settings, workspace, backend, store)
    mapping = {
        "full_name": {"columns": [{"index": 0, "header": "Name"}]},
        "city": {"columns": [{"index": 1, "header": "City"}]},
        "state": {"columns": [{"index": 2, "header": "State"}]},
        "firm_name": {"columns": [{"index": 3, "header": "Firm Name"}]},
    }
    async with backend.run_scope(run_id, corp_id, conversation_id):
        validation = await tools["validate_advisor_input"].ainvoke(
            {"attachment_id": derived.attachment_id, "mapping": mapping}
        )
        result = await tools["create_advisor_match"].ainvoke(
            {
                "attachment_id": derived.attachment_id,
                "mapping": mapping,
                "mapping_fingerprint": validation["mapping_fingerprint"],
            }
        )
        page = await tools["list_advisor_match_results"].ainvoke(
            {
                "match_session_id": result["match_session_id"],
                "status": "ambiguous_match",
            }
        )
        changed = await tools["apply_advisor_match_decisions"].ainvoke(
            {
                "match_session_id": result["match_session_id"],
                "decisions": [
                    {
                        "review_item_id": page["items"][0]["review_item_id"],
                        "action": "confirm_candidate",
                        "crd_number": page["items"][0]["candidates"][0][
                            "crd_number"
                        ],
                    }
                ],
            }
        )
    assert result["source_transformation"]["type"] == "bulk_firm_augmentation"
    assert changed["counts"]["matched"] == 1
    assert changed["output_artifact_id"] != result["output_artifact_id"]
    store.close()


@pytest.mark.asyncio
async def test_attachment_id_cannot_cross_conversations(settings) -> None:
    corp_id = "A123456"
    workspace = Workspace(settings.data_root)
    store = Store(settings.application_db, settings.data_root)
    first = store.create_conversation(corp_id=corp_id)
    first_run_id, _ = store.create_run(
        first.conversation_id, "upload", corp_id=corp_id
    )
    attachment_id = _upload(
        workspace,
        store,
        corp_id=corp_id,
        conversation_id=first.conversation_id,
        run_id=first_run_id,
        name="private.csv",
        content=b"Name\nPrivate Person\n",
    )
    store.finish_run(first_run_id, "completed", corp_id=corp_id)

    second = store.create_conversation(corp_id=corp_id)
    second_run_id, _ = store.create_run(
        second.conversation_id, "inspect", corp_id=corp_id
    )
    backend = AdvisorWorkspaceBackend(settings.runtime_root)
    tools = _tools(settings, workspace, backend, store)
    async with backend.run_scope(second_run_id, corp_id, second.conversation_id):
        with pytest.raises(ValueError, match="No validated advisor input"):
            await tools["get_current_advisor_input"].ainvoke({})
        with pytest.raises(ValueError, match="unknown in the current chat"):
            await tools["inspect_advisor_upload"].ainvoke(
                {"attachment_id": attachment_id}
            )
        with pytest.raises(ValueError, match="unknown in the current chat"):
            await tools["create_advisor_match"].ainvoke(
                {
                    "attachment_id": attachment_id,
                    "mapping": {
                        "full_name": {
                            "columns": [{"index": 0, "header": "Name"}]
                        }
                    },
                    "mapping_fingerprint": "0" * 64,
                }
            )
    store.close()


@pytest.mark.asyncio
async def test_unlisted_crd_requires_proposal_then_later_turn_confirmation(
    settings,
) -> None:
    corp_id = "A123456"
    workspace = Workspace(settings.data_root)
    store = Store(settings.application_db, settings.data_root)
    conversation_id = store.create_conversation(corp_id=corp_id).conversation_id
    first_run_id, _ = store.create_run(
        conversation_id, "match", corp_id=corp_id
    )
    attachment_id = _upload(
        workspace,
        store,
        corp_id=corp_id,
        conversation_id=conversation_id,
        run_id=first_run_id,
        name="unknown.csv",
        content=b"Name\nUnknown Person\n",
    )
    backend = AdvisorWorkspaceBackend(settings.runtime_root)
    tools = _tools(settings, workspace, backend, store)
    mapping = {"full_name": {"columns": [{"index": 0, "header": "Name"}]}}
    async with backend.run_scope(first_run_id, corp_id, conversation_id):
        validation = await tools["validate_advisor_input"].ainvoke(
            {"attachment_id": attachment_id, "mapping": mapping}
        )
        result = await tools["create_advisor_match"].ainvoke(
            {
                "attachment_id": attachment_id,
                "mapping": mapping,
                "mapping_fingerprint": validation["mapping_fingerprint"],
                "firm_resolution": "continue_without_firm",
            }
        )
        page = await tools["list_advisor_match_results"].ainvoke(
            {"match_session_id": result["match_session_id"], "status": "no_match"}
        )
        assert page["total"] == 1
        unmatched = await tools["list_advisor_match_results"].ainvoke(
            {"match_session_id": result["match_session_id"], "status": "unmatched"}
        )
        assert unmatched["items"] == page["items"]
        proposal = await tools["propose_crd_match"].ainvoke(
            {
                "match_session_id": result["match_session_id"],
                "review_item_id": page["items"][0]["review_item_id"],
                "crd_number": "  99000001  ",
            }
        )
        assert proposal["requires_explicit_confirmation"] is True
        assert proposal["resolved_advisor"]["crd_number"] == "99000001"
        with pytest.raises(ValueError, match="later user turn"):
            await tools["apply_advisor_match_decisions"].ainvoke(
                {
                    "match_session_id": result["match_session_id"],
                    "decisions": [
                        {
                            "review_item_id": page["items"][0]["review_item_id"],
                            "action": "confirm_manual_crd",
                            "proposal_id": proposal["proposal_id"],
                        }
                    ],
                }
            )
    store.finish_run(first_run_id, "completed", corp_id=corp_id)
    confirm_run_id, _ = store.create_run(
        conversation_id, "confirm", corp_id=corp_id
    )
    async with backend.run_scope(confirm_run_id, corp_id, conversation_id):
        confirmed = await tools["apply_advisor_match_decisions"].ainvoke(
            {
                "match_session_id": result["match_session_id"],
                "decisions": [
                    {
                        "review_item_id": page["items"][0]["review_item_id"],
                        "action": "confirm_manual_crd",
                        "proposal_id": proposal["proposal_id"],
                    }
                ],
            }
        )
        assert confirmed["counts"]["matched"] == 1
    store.close()
