from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
import yaml
from openpyxl import Workbook, load_workbook

from advisor_match.advisor_matching import normalization as norm
from advisor_match.advisor_matching import policy
from advisor_match.advisor_matching.input_loader import validate_and_load_input
from advisor_match.advisor_matching.index import AdvisorIndex, ReferenceDataQualityError
from advisor_match.advisor_matching.matcher import run_matching
from advisor_match.advisor_matching.schemas import (
    AdvisorRecord,
    InputMapping,
    MatchDecision,
)
from advisor_match.advisor_matching.source import SyntheticAdvisorReferenceSource
from advisor_match.files import InMemoryFile


ADVISORS = [
    AdvisorRecord(
        crd_number="1001",
        first_name="Robert",
        last_name="Mercer",
        firm_name="Cedar Grove Advisory LLC",
        email="robert@example.com",
        city="Richmond",
        state="VA",
        zip_code="23219",
    ),
    AdvisorRecord(
        crd_number="1002",
        first_name="John",
        last_name="Smith",
        firm_name="Northstar Wealth",
        email="shared@example.com",
        city="Boston",
        state="MA",
        zip_code="02108",
    ),
    AdvisorRecord(
        crd_number="1003",
        first_name="John",
        last_name="Smith",
        firm_name="Harbor Advisory",
        email="shared@example.com",
        city="Cambridge",
        state="MA",
        zip_code="02139",
    ),
]


def _row(row_number: int, **mapped: str):
    values = {
        "crd_number": "",
        "first_name": "",
        "last_name": "",
        "full_name": "",
        "firm_name": "",
        "email": "",
        "city": "",
        "state": "",
        "zip_code": "",
    }
    values.update(mapped)
    return row_number, dict(values), values


def _binding(index: int, header: str | None) -> dict:
    return {"index": index, "header": header}


def _memory_file(path: Path) -> InMemoryFile:
    return InMemoryFile(path.name, path.read_bytes())


def test_firm_normalization_removes_legal_suffix_variations() -> None:
    assert norm.firm("Morgan Stanley") == norm.firm("Morgan Stanley, LLC")
    assert norm.firm("Morgan Stanley") == norm.firm("Morgan Stanley & Co., LLC")


def test_crd_is_an_opaque_trimmed_identifier() -> None:
    assert norm.crd("  FSA_ID:111  ") == "FSA_ID:111"
    assert norm.crd("1001.0") == "1001.0"
    assert norm.crd("  ") == ""


def test_exact_opaque_crd_is_decisive() -> None:
    advisors = [
        AdvisorRecord(
            crd_number="FSA_ID:111",
            first_name="Jane",
            last_name="Doe",
            firm_name="",
            email="",
            city="",
            state="",
            zip_code="",
        )
    ]

    decisions, counts, _ = run_matching(
        [_row(2, crd_number="  FSA_ID:111  ")], advisors
    )

    assert counts.matched == 1
    assert decisions[0].rule_id == "EXACT_CRD"
    assert decisions[0].matched_advisor.crd_number == "FSA_ID:111"


def test_numeric_looking_crd_is_not_transformed() -> None:
    decisions, counts, _ = run_matching(
        [_row(2, crd_number="1001.0")], ADVISORS
    )

    assert counts.no_match == 1
    assert decisions[0].rule_id == "CRD_NOT_FOUND"


def test_synthetic_source_accepts_opaque_crd(tmp_path) -> None:
    source = tmp_path / "opaque-crd.csv"
    source.write_text(
        "CRD_NUMBER,FIRST_NAME,LAST_NAME,FIRM_NAME,EMAIL,CITY,STATE,ZIP_CODE\n"
        " FSA_ID:111 ,Jane,Doe,,,,,\n",
        encoding="utf-8",
    )

    records = list(SyntheticAdvisorReferenceSource(source).iter_records())

    assert records[0].crd_number == "FSA_ID:111"


def test_synthetic_source_rejects_duplicate_trimmed_crds(tmp_path) -> None:
    source = tmp_path / "duplicate-opaque-crd.csv"
    source.write_text(
        "CRD_NUMBER,FIRST_NAME,LAST_NAME,FIRM_NAME,EMAIL,CITY,STATE,ZIP_CODE\n"
        " FSA_ID:111 ,Jane,Doe,,,,,\n"
        "FSA_ID:111,Janet,Doe,,,,,\n",
        encoding="utf-8",
    )

    with pytest.raises(ReferenceDataQualityError) as raised:
        AdvisorIndex.from_records(
            SyntheticAdvisorReferenceSource(source).iter_records()
        )

    assert raised.value.duplicate_crds == {"FSA_ID:111": 2}


def test_exact_crd_is_decisive_and_reports_conflicts() -> None:
    decisions, counts, _ = run_matching(
        [
            _row(
                2,
                crd_number="1001",
                full_name="Someone Else",
                firm_name="Different Firm",
                state="CA",
                email="other@example.com",
            )
        ],
        ADVISORS,
    )
    assert counts.matched == 1
    assert decisions[0].rule_id == "EXACT_CRD"
    assert decisions[0].matched_advisor.crd_number == "1001"
    assert any("email conflicted" in warning for warning in decisions[0].warnings)
    assert any("Firm conflicts" in warning for warning in decisions[0].warnings)


def test_unique_email_can_match_with_unknown_crd_warning() -> None:
    decisions, counts, _ = run_matching(
        [_row(2, crd_number="9999", email="ROBERT@EXAMPLE.COM")], ADVISORS
    )
    assert counts.matched == 1
    assert decisions[0].rule_id == "UNIQUE_EXACT_EMAIL"
    assert any("CRD was not found" in warning for warning in decisions[0].warnings)


def test_non_unique_email_is_ambiguous_even_with_other_evidence() -> None:
    decisions, counts, _ = run_matching(
        [
            _row(
                2,
                email="SHARED@example.com",
                full_name="John Smith",
                firm_name="Northstar Wealth",
            )
        ],
        ADVISORS,
    )
    assert counts.ambiguous_match == 1
    assert decisions[0].rule_id == "NON_UNIQUE_EMAIL"
    assert {item.crd_number for item in decisions[0].candidates} == {"1002", "1003"}
    assert decisions[0].candidate_count == 2
    assert decisions[0].candidates_truncated is False


@pytest.mark.parametrize(
    ("mapped", "rule_id"),
    [
        ({"crd_number": "12-34"}, "CRD_NOT_FOUND"),
        ({"crd_number": "9999"}, "CRD_NOT_FOUND"),
        ({"email": "not-an-email"}, "MALFORMED_EMAIL"),
        ({"full_name": "Smith"}, "INSUFFICIENT_NAME"),
    ],
)
def test_invalid_identifier_rows_have_specific_reasons(mapped, rule_id) -> None:
    decisions, counts, _ = run_matching([_row(2, **mapped)], ADVISORS)
    assert counts.no_match == 1
    assert decisions[0].rule_id == rule_id


def test_malformed_value_is_warning_when_other_evidence_matches() -> None:
    decisions, counts, _ = run_matching(
        [
            _row(
                2,
                email="not-an-email",
                full_name="Robert Mercer",
                firm_name="Cedar Grove Advisory",
            )
        ],
        ADVISORS,
    )
    assert counts.matched == 1
    assert decisions[0].rule_id == "EXACT_NAME_SUPPORTED"
    assert "The supplied email is malformed." in decisions[0].warnings


def test_nickname_with_support_is_review_only() -> None:
    decisions, counts, _ = run_matching(
        [
            _row(
                2,
                full_name="Bob Mercer",
                firm_name="Cedar Grove Advisory",
                city="Richmond",
                state="VA",
            )
        ],
        ADVISORS,
    )
    assert counts.ambiguous_match == 1
    assert decisions[0].rule_id == "NICKNAME_REVIEW_REQUIRED"
    assert decisions[0].candidates[0].crd_number == "1001"
    assert any(
        "Nickname alias" in evidence
        or "Curated nickname" in evidence
        for evidence in decisions[0].candidates[0].supporting_evidence
    )


def test_name_alone_is_never_confirmed() -> None:
    decisions, counts, _ = run_matching(
        [_row(2, full_name="John Smith")], ADVISORS
    )
    assert counts.matched == 0
    assert decisions[0].status == "Ambiguous Match"
    assert decisions[0].rule_id == "EXACT_NAME_REVIEW_REQUIRED"
    assert len(decisions[0].candidates) == 2


def test_legal_suffix_normalization_supports_exact_name() -> None:
    decisions, counts, _ = run_matching(
        [
            _row(
                2,
                full_name="Mercer, Robert",
                firm_name="Cedar Grove Advisory",
            )
        ],
        ADVISORS,
    )
    assert counts.matched == 1
    assert decisions[0].matched_advisor.crd_number == "1001"
    assert "Exact normalized firm" in decisions[0].matched_advisor.supporting_evidence


def test_close_firm_can_support_exact_name() -> None:
    decisions, counts, _ = run_matching(
        [
            _row(
                2,
                full_name="Robert Mercer",
                firm_name="Cedar Grove Advisry",
            )
        ],
        ADVISORS,
    )
    assert counts.matched == 1
    assert any(
        "Close firm name" in evidence
        for evidence in decisions[0].matched_advisor.supporting_evidence
    )


@pytest.mark.parametrize(
    ("input_firm", "master_firm"),
    [
        ("Ed Jones", "Edward D. Jones Financial LLC"),
        ("North Star", "Northstar Financial Group"),
    ],
)
def test_anchored_firm_wildcard_can_support_exact_name(
    input_firm: str, master_firm: str
) -> None:
    advisors = [
        AdvisorRecord(
            crd_number="2001",
            first_name="Alice",
            last_name="Brown",
            firm_name=master_firm,
        )
    ]
    decisions, counts, _ = run_matching(
        [_row(2, first_name="Alice", last_name="Brown", firm_name=input_firm)],
        advisors,
    )

    assert counts.matched == 1
    assert decisions[0].rule_id == "EXACT_NAME_SUPPORTED"
    assert any(
        "Anchored firm wildcard matched" in evidence
        for evidence in decisions[0].matched_advisor.supporting_evidence
    )


def test_generic_firm_wildcard_can_support_exact_name() -> None:
    advisors = [
        AdvisorRecord(
            crd_number="2001",
            first_name="Alice",
            last_name="Brown",
            firm_name="Financial Group Advisors",
        )
    ]
    decisions, counts, _ = run_matching(
        [
            _row(
                2,
                first_name="Alice",
                last_name="Brown",
                firm_name="Financial Group",
            )
        ],
        advisors,
    )

    assert counts.matched == 1
    assert decisions[0].rule_id == "EXACT_NAME_SUPPORTED"


def test_short_firm_wildcard_cannot_support_exact_name() -> None:
    advisors = [
        AdvisorRecord(
            crd_number="2001",
            first_name="Alice",
            last_name="Brown",
            firm_name="Edward Financial",
        )
    ]
    decisions, counts, _ = run_matching(
        [_row(2, first_name="Alice", last_name="Brown", firm_name="Ed")],
        advisors,
    )

    assert counts.matched == 0
    assert decisions[0].status == "Ambiguous Match"


def test_firm_wildcard_does_not_override_state_conflict() -> None:
    advisors = [
        AdvisorRecord(
            crd_number="2001",
            first_name="Alice",
            last_name="Brown",
            firm_name="Edward D. Jones Financial LLC",
            state="VA",
        )
    ]
    decisions, counts, _ = run_matching(
        [
            _row(
                2,
                first_name="Alice",
                last_name="Brown",
                firm_name="Ed Jones",
                state="CA",
            )
        ],
        advisors,
    )

    assert counts.matched == 0
    assert decisions[0].status == "Ambiguous Match"


def test_firm_wildcard_does_not_choose_between_duplicate_names() -> None:
    advisors = [
        AdvisorRecord(
            crd_number="2001",
            first_name="Alice",
            last_name="Brown",
            firm_name="Edward D. Jones Financial LLC",
        ),
        AdvisorRecord(
            crd_number="2002",
            first_name="Alice",
            last_name="Brown",
            firm_name="Edwin Jones Wealth LLC",
        ),
    ]
    decisions, counts, _ = run_matching(
        [
            _row(
                2,
                first_name="Alice",
                last_name="Brown",
                firm_name="Ed Jones",
            )
        ],
        advisors,
    )

    assert counts.ambiguous_match == 1
    assert {candidate.crd_number for candidate in decisions[0].candidates} == {
        "2001",
        "2002",
    }


def test_name_typo_is_no_match_even_with_firm_or_location() -> None:
    without_location, counts, _ = run_matching(
        [
            _row(
                2,
                full_name="Robrt Mercer",
                firm_name="Cedar Grove Advisry",
            )
        ],
        ADVISORS,
    )
    assert counts.no_match == 1
    assert without_location[0].rule_id == "NAME_NOT_FOUND"
    with_location, counts, _ = run_matching(
        [
            _row(
                2,
                full_name="Robrt Mercer",
                firm_name="Cedar Grove Advisry",
                city="Richmond",
                state="VA",
            )
        ],
        ADVISORS,
    )
    assert counts.no_match == 1
    assert with_location[0].rule_id == "NAME_NOT_FOUND"


def test_middle_name_is_ignored_for_exact_index_lookup() -> None:
    decisions, counts, _ = run_matching(
        [
            _row(
                2,
                full_name="Robert Allen Mercer",
                firm_name="Cedar Grove Advisory",
            )
        ],
        ADVISORS,
    )
    assert counts.matched == 1
    assert decisions[0].rule_id == "EXACT_NAME_SUPPORTED"


def test_uncommaed_compound_surname_limitation_is_explicit() -> None:
    advisors = [
        AdvisorRecord(
            crd_number="2001",
            first_name="Maria",
            last_name="De La Cruz",
            firm_name="Harbor Advisory",
        )
    ]
    decisions, counts, _ = run_matching(
        [_row(2, full_name="Maria De La Cruz", firm_name="Harbor Advisory")],
        advisors,
    )
    assert counts.no_match == 1
    assert decisions[0].rule_id == "NAME_NOT_FOUND"

    separated, separated_counts, _ = run_matching(
        [
            _row(
                3,
                first_name="Maria",
                last_name="De La Cruz",
                firm_name="Harbor Advisory",
            )
        ],
        advisors,
    )
    assert separated_counts.matched == 1
    assert separated[0].rule_id == "EXACT_NAME_SUPPORTED"


def test_candidate_pool_is_bounded_and_reports_total() -> None:
    advisors = [
        AdvisorRecord(
            crd_number=str(3000 + index),
            first_name="John",
            last_name="Smith",
        )
        for index in range(5)
    ]
    decisions, counts, _ = run_matching(
        [_row(2, full_name="John Smith")], advisors
    )
    assert counts.ambiguous_match == 1
    assert len(decisions[0].candidates) == 3
    assert decisions[0].candidate_count == 5
    assert decisions[0].candidates_truncated is True


def test_candidate_order_uses_explicit_evidence_precedence() -> None:
    advisors = [
        AdvisorRecord(
            crd_number="3003",
            first_name="John",
            last_name="Smith",
            firm_name="Northstar Global Wealth Partners",
        ),
        AdvisorRecord(
            crd_number="3002",
            first_name="John",
            last_name="Smith",
            city="Boston",
            state="MA",
        ),
        AdvisorRecord(
            crd_number="3001",
            first_name="John",
            last_name="Smith",
            firm_name="Northstar Wealth",
        ),
    ]
    decisions, counts, _ = run_matching(
        [
            _row(
                2,
                full_name="John Smith",
                firm_name="Northstar Wealth",
                city="Boston",
                state="MA",
            )
        ],
        advisors,
    )
    assert counts.ambiguous_match == 1
    assert [item.crd_number for item in decisions[0].candidates] == [
        "3001",
        "3002",
        "3003",
    ]


def test_match_decision_backfills_candidate_count() -> None:
    legacy = MatchDecision.model_validate(
        {
            "review_item_id": "ami_legacy",
            "source_row_number": 2,
            "source_values": {"Name": "John Smith"},
            "mapped_values": {"full_name": "John Smith"},
            "status": "Ambiguous Match",
            "confidence": "Uncertain",
            "rule_id": "AMBIGUOUS_CANDIDATES",
            "explanation": "Legacy decision",
            "candidates": [ADVISORS[1].model_dump()],
        }
    )
    assert legacy.candidate_count == 1
    assert legacy.candidates_truncated is False


def test_firm_and_state_conflicts_block_name_based_auto_match() -> None:
    for mapped in (
        {
            "full_name": "Robert Mercer",
            "firm_name": "Different Company",
            "city": "Richmond",
            "state": "VA",
        },
        {
            "full_name": "Robert Mercer",
            "firm_name": "Cedar Grove Advisory",
            "city": "Richmond",
            "state": "CA",
        },
    ):
        decisions, counts, _ = run_matching([_row(2, **mapped)], ADVISORS)
        assert counts.ambiguous_match == 1
        assert decisions[0].status == "Ambiguous Match"


def test_same_state_city_difference_is_weak_when_firm_is_exact() -> None:
    decisions, counts, _ = run_matching(
        [
            _row(
                2,
                full_name="Robert Mercer",
                firm_name="Cedar Grove Advisory",
                city="Norfolk",
                state="VA",
            )
        ],
        ADVISORS,
    )
    assert counts.matched == 1
    assert any(
        "City differs" in evidence
        for evidence in decisions[0].matched_advisor.conflicting_evidence
    )


def test_zip_is_context_only_and_cannot_support_name() -> None:
    decisions, counts, _ = run_matching(
        [_row(2, full_name="Robert Mercer", zip_code="23219")], ADVISORS
    )
    assert counts.ambiguous_match == 1
    assert decisions[0].candidates[0].contextual_evidence == ["ZIP matches"]


def test_duplicate_input_rows_remain_separate_review_items() -> None:
    rows = [_row(2, crd_number="1001"), _row(3, crd_number="1001")]
    decisions, counts, warnings = run_matching(rows, ADVISORS)
    assert counts.matched == 2
    assert decisions[0].review_item_id != decisions[1].review_item_id
    assert decisions[0].duplicate_group == decisions[1].duplicate_group
    assert warnings


def test_mapping_requires_complete_name_or_strong_identifier() -> None:
    with pytest.raises(ValueError, match="both first and last"):
        InputMapping(last_name=_binding(0, "Surname"))
    with pytest.raises(ValueError, match="Extra inputs are not permitted"):
        InputMapping.model_validate(
            {
                "full_name": _binding(0, "Name"),
                "street_address": _binding(1, "Address"),
            }
        )


def test_loader_supports_later_headers_and_skips_blank_rows(tmp_path) -> None:
    source = tmp_path / "later.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Advisors"
    sheet.append(["Quarterly export"])
    sheet.append([])
    sheet.append(["Name", "Firm"])
    sheet.append(["Robert Mercer", "Cedar Grove Advisory"])
    sheet.append([])
    sheet.append(["John Smith", "Northstar Wealth"])
    workbook.save(source)
    mapping = InputMapping(
        sheet_name="Advisors",
        header_row=3,
        full_name=_binding(0, "Name"),
        firm_name=_binding(1, "Firm"),
    )
    loaded = validate_and_load_input(_memory_file(source), mapping, max_rows=100)
    assert [row[0] for row in loaded.rows] == [4, 6]
    assert loaded.summary.data_row_count == 2
    assert loaded.summary.blank_row_count == 1
    assert loaded.summary.preamble_row_count == 2


def test_loader_supports_headerless_input(tmp_path) -> None:
    source = tmp_path / "headerless.csv"
    source.write_text(
        "1001,Robert,Mercer,Cedar Grove Advisory\n"
        "1002,John,Smith,Northstar Wealth\n",
        encoding="utf-8",
    )
    mapping = InputMapping(
        header_row=None,
        crd_number=_binding(0, None),
        first_name=_binding(1, None),
        last_name=_binding(2, None),
        firm_name=_binding(3, None),
    )
    loaded = validate_and_load_input(_memory_file(source), mapping, max_rows=100)
    assert loaded.rows[0][0] == 1
    assert loaded.rows[0][1]["Column A"] == "1001"
    assert loaded.columns[0] == {"index": 0, "header": None, "label": "Column A"}


def test_loader_selects_exact_sheet_header_and_duplicate_column_index() -> None:
    workbook = Workbook()
    workbook.active.title = "Ignore"
    workbook.active.append(["Not", "Advisors"])
    advisors = workbook.create_sheet("Advisors")
    advisors.append(["Quarterly export"])
    advisors.append(["CRD", "CRD", "Name"])
    advisors.append(["wrong", "1001", "Avery Stone"])
    output = BytesIO()
    workbook.save(output)
    source = InMemoryFile("multi.xlsx", output.getvalue())
    mapping = InputMapping(
        sheet_name="Advisors",
        header_row=2,
        crd_number=_binding(1, "CRD"),
        full_name=_binding(2, "Name"),
    )

    loaded = validate_and_load_input(source, mapping, max_rows=10)

    assert loaded.selected_sheet == "Advisors"
    assert loaded.rows[0][2]["crd_number"] == "1001"
    assert [column["label"] for column in loaded.columns] == [
        "CRD",
        "CRD [2]",
        "Name",
    ]


def test_loader_rejects_changed_exact_header(tmp_path) -> None:
    source = tmp_path / "changed.csv"
    source.write_text("Advisor\nRobert Mercer\n", encoding="utf-8")
    mapping = InputMapping(full_name=_binding(0, "Name"))
    with pytest.raises(ValueError, match="expected 'Name', observed 'Advisor'"):
        validate_and_load_input(_memory_file(source), mapping, max_rows=100)


def test_missing_firm_evidence_is_row_based(tmp_path) -> None:
    source = tmp_path / "missing-firm.csv"
    source.write_text(
        "Name,Firm,CRD,Email\n"
        "Robert Mercer,,,\n"
        "John Smith,Northstar Wealth,,\n"
        "Avery Stone,,,avery@example.com\n",
        encoding="utf-8",
    )
    mapping = InputMapping(
        full_name=_binding(0, "Name"),
        firm_name=_binding(1, "Firm"),
        crd_number=_binding(2, "CRD"),
        email=_binding(3, "Email"),
    )
    loaded = validate_and_load_input(_memory_file(source), mapping, max_rows=100)
    assert loaded.summary.missing_firm_row_count == 1
    assert loaded.missing_firm_sample[0]["source_row_number"] == 2


def test_missing_firm_column_does_not_require_confirmation_with_email(tmp_path) -> None:
    source = tmp_path / "email-only.csv"
    source.write_text(
        "Name,Email\nRobert Mercer,robert@example.com\n",
        encoding="utf-8",
    )
    mapping = InputMapping(
        full_name=_binding(0, "Name"),
        email=_binding(1, "Email"),
    )

    loaded = validate_and_load_input(_memory_file(source), mapping, max_rows=100)

    assert loaded.summary.firm_column_missing is True
    assert loaded.summary.missing_firm_row_count == 0
    assert loaded.summary.missing_firm_confirmation_required is False


def test_input_loader_stops_at_configured_row_limit(tmp_path) -> None:
    source = tmp_path / "too-many.csv"
    source.write_text("Name\nOne Person\nTwo Person\n", encoding="utf-8")
    mapping = InputMapping(full_name=_binding(0, "Name"))
    with pytest.raises(ValueError, match="limit is 1"):
        validate_and_load_input(_memory_file(source), mapping, max_rows=1)


def test_blank_rows_cannot_hide_records_beyond_row_limit(tmp_path) -> None:
    source = tmp_path / "blank-gap.csv"
    source.write_text(
        "Name\nOne Person\n\nTwo Person\n",
        encoding="utf-8",
    )
    mapping = InputMapping(full_name=_binding(0, "Name"))
    with pytest.raises(ValueError, match="limit is 1"):
        validate_and_load_input(_memory_file(source), mapping, max_rows=1)


def test_unreadable_xlsx_is_reported_as_a_structural_input_error(tmp_path) -> None:
    source = tmp_path / "broken.xlsx"
    source.write_bytes(b"not an Excel workbook")
    mapping = InputMapping(full_name=_binding(0, "Name"))
    with pytest.raises(ValueError, match="could not be read"):
        validate_and_load_input(_memory_file(source), mapping, max_rows=100)


@pytest.mark.parametrize(
    ("filename", "mapping", "expected"),
    [
        (
            "clean_exact_matches.csv",
            {
                "crd_number": _binding(0, "CRD_NUMBER"),
                "first_name": _binding(1, "FIRST_NAME"),
                "last_name": _binding(2, "LAST_NAME"),
                "email": _binding(3, "EMAIL"),
            },
            (3, 0, 0),
        ),
        (
            "casing_and_whitespace.xlsx",
            {
                "sheet_name": "Advisors",
                "first_name": _binding(0, "first name"),
                "last_name": _binding(1, "last name"),
                "email": _binding(2, "email address"),
                "firm_name": _binding(3, "firm"),
            },
            (2, 0, 0),
        ),
        (
            "location_variations.csv",
            {
                "first_name": _binding(0, "First Name"),
                "last_name": _binding(1, "Last Name"),
                "firm_name": _binding(2, "Firm Name"),
                "city": _binding(3, "City"),
                "state": _binding(4, "State"),
                "zip_code": _binding(5, "ZIP"),
            },
            (1, 1, 0),
        ),
        (
            "missing_fields.xlsx",
            {
                "sheet_name": "Advisors",
                "first_name": _binding(0, "First Name"),
                "last_name": _binding(1, "Last Name"),
                "email": _binding(2, "Email"),
                "firm_name": _binding(3, "Firm Name"),
                "city": _binding(4, "City"),
                "state": _binding(5, "State"),
            },
            (1, 1, 1),
        ),
        (
            "partial_matches.csv",
            {
                "full_name": _binding(0, "Advisor Name"),
                "firm_name": _binding(1, "Company"),
                "city": _binding(2, "City"),
                "state": _binding(3, "State"),
            },
            (0, 2, 1),
        ),
        (
            "duplicate_rows.xlsx",
            {
                "sheet_name": "Advisors",
                "crd_number": _binding(0, "CRD"),
                "first_name": _binding(1, "First Name"),
                "last_name": _binding(2, "Last Name"),
                "email": _binding(3, "Email"),
            },
            (3, 0, 0),
        ),
        (
            "unknown_advisors.csv",
            {
                "first_name": _binding(0, "First Name"),
                "last_name": _binding(1, "Last Name"),
                "firm_name": _binding(2, "Firm Name"),
                "city": _binding(3, "City"),
                "state": _binding(4, "State"),
            },
            (0, 0, 2),
        ),
    ],
)
def test_generated_fixtures_have_expected_results(filename, mapping, expected) -> None:
    root = Path(__file__).parents[1]
    advisors = list(
        SyntheticAdvisorReferenceSource(
            root
            / "advisor_match"
            / "advisor_matching"
            / "data"
            / "master_advisors.csv"
        ).iter_records()
    )
    loaded = validate_and_load_input(
        _memory_file(root / "examples" / "advisor-match" / filename),
        InputMapping.model_validate(mapping),
        max_rows=100,
    )
    _, counts, _ = run_matching(loaded.rows, advisors)
    assert (counts.matched, counts.ambiguous_match, counts.no_match) == expected


def test_skill_policy_reference_matches_executable_policy() -> None:
    root = Path(__file__).parents[1]
    documented = yaml.safe_load(
        (
            root
            / "docs"
            / "contracts"
            / "matching-policy.yaml"
        ).read_text(encoding="utf-8")
    )
    assert documented["version"] == policy.POLICY_VERSION
    firm = documented["firm_similarity"]
    assert firm["minimum_similarity"] == policy.MINIMUM_FIRM_SIMILARITY
    assert firm["conflict_similarity"] == policy.FIRM_CONFLICT_SIMILARITY
    wildcard = documented["firm_wildcard"]
    assert wildcard["anchored"] is True
    assert wildcard["minimum_length"] == policy.FIRM_WILDCARD_MIN_LENGTH
