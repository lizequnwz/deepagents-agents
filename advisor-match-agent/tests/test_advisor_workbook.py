from __future__ import annotations

from datetime import UTC, datetime

from openpyxl import load_workbook

from general_agent.advisor_matching.schemas import (
    InputMapping,
    InputSummary,
    MatchCounts,
    MatchDecision,
    ReferenceSnapshotManifest,
)
from general_agent.advisor_matching.workbook import (
    SHEETS,
    verify_match_workbook,
    write_match_workbook,
)


def test_workbook_is_human_first_styled_auditable_and_formula_safe(tmp_path) -> None:
    decision = MatchDecision(
        review_item_id="ami_test",
        source_row_number=2,
        source_values={"Name": '=HYPERLINK("https://invalid.example")'},
        mapped_values={
            "full_name": '=HYPERLINK("https://invalid.example")',
            "city": "Boston",
            "state": "MA",
            "zip_code": "02108",
        },
        status="No Match",
        confidence="None",
        rule_id="NO_ACCEPTABLE_CANDIDATE",
        explanation="No advisor satisfied the accepted identity rules.",
    )
    reference = ReferenceSnapshotManifest(
        reference_snapshot_id="ars_" + "a" * 32,
        row_count=40,
        columns=["CRD_NUMBER"],
        source_kind="synthetic",
        schema_version="1",
        retrieved_at=datetime.now(UTC),
        sha256="a" * 64,
    )
    output = tmp_path / "advisor_matches.xlsx"
    write_match_workbook(
        output,
        session_id="ams_test",
        decisions=[decision],
        counts=MatchCounts(no_match=1),
        mapping=InputMapping(
            full_name={"columns": [{"index": 0, "header": "Name"}]}
        ),
        input_summary=InputSummary(
            data_row_count=1,
            blank_row_count=0,
            preamble_row_count=0,
            missing_firm_row_count=1,
            missing_firm_confirmation_required=True,
        ),
        source_name="input.csv",
        source_sha256="b" * 64,
        reference=reference,
        policy_version="2",
    )
    assert verify_match_workbook(output, expected_rows=1) == {
        "matched": 0,
        "ambiguous_match": 0,
        "no_match": 1,
        "review_items": 1,
        "original": 1,
    }

    workbook = load_workbook(output, data_only=False)
    assert tuple(workbook.sheetnames) == SHEETS
    original_value = workbook["Original Input"]["B2"]
    assert original_value.value.startswith("=")
    assert original_value.data_type == "s"
    for sheet in workbook.worksheets:
        assert sheet.freeze_panes == "A2"
        assert sheet.auto_filter.ref
        assert sheet.sheet_view.showGridLines is False

    matched = workbook["Matched"]
    review = workbook["Review Required"]
    matched_headers = [cell.value for cell in matched[1]]
    review_headers = [cell.value for cell in review[1]]
    visible_matched = [
        header
        for index, header in enumerate(matched_headers, start=1)
        if not matched.column_dimensions[matched.cell(1, index).column_letter].hidden
    ]
    visible_review = [
        header
        for index, header in enumerate(review_headers, start=1)
        if not review.column_dimensions[review.cell(1, index).column_letter].hidden
    ]
    assert len(visible_matched) == 17
    assert len(visible_review) == 23
    assert "Candidate Pool Size" in visible_review
    assert "Candidates Truncated" in visible_review
    assert "User Decision" in visible_review
    assert "Selected CRD" in visible_review
    assert "Reviewer Notes" in visible_review
    assert "Review Item ID" not in visible_review
    assert matched.column_dimensions["A"].width >= len("Source Row")
    assert matched.row_dimensions[1].height == 24
    assert review.row_dimensions[2].height == 36
    assert matched["A1"].fill.fgColor.rgb.endswith("1F4E78")
    status_column = review_headers.index("Status") + 1
    assert review.cell(2, status_column).fill.fgColor.rgb.endswith("F4CCCC")
    user_decision_column = review_headers.index("User Decision") + 1
    assert review.cell(1, user_decision_column).fill.fgColor.rgb.endswith("FFF2CC")
    assert review.cell(2, user_decision_column).fill.fgColor.rgb.endswith("FFF2CC")
    summary = {
        row[0].value: row[1].value
        for row in workbook["Run Summary"].iter_rows(min_row=2)
    }
    assert "Review Required" in summary["Manual Review"]
    assert "not sent back to or validated" in summary["Local Edit Boundary"]
    assert summary["Session Status"] == "Matching Complete"
    workbook.close()
