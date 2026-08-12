"""Validate and load mapped source rows entirely from request memory."""

from __future__ import annotations

import csv
import hashlib
import json
from dataclasses import dataclass
from io import StringIO
from zipfile import BadZipFile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from advisor_match.advisor_matching import normalization as norm
from advisor_match.advisor_matching.schemas import InputMapping, InputSummary, MAPPED_FIELDS
from advisor_match.files import InMemoryFile, InvalidUploadError


@dataclass(frozen=True, slots=True)
class LoadedInput:
    rows: list[tuple[int, dict[str, object], dict[str, str]]]
    selected_sheet: str | None
    columns: list[dict[str, object]]
    source_sha256: str
    mapping_fingerprint: str
    summary: InputSummary
    missing_firm_sample: list[dict[str, object]]


def validate_and_load_input(
    source: InMemoryFile,
    mapping: InputMapping,
    *,
    max_rows: int,
    missing_firm_sample_limit: int = 3,
) -> LoadedInput:
    """Validate exact sheet/header/column bindings and load bounded rows."""

    source.validate_table_type()
    selected_sheet = _selected_sheet(source, mapping)
    data_start = mapping.header_row or 0
    read_limit = data_start + max_rows + 1
    raw = _read_raw(source, sheet_name=selected_sheet, nrows=read_limit)
    if mapping.header_row is not None and mapping.header_row > len(raw.index):
        raise ValueError(
            f"Header row {mapping.header_row} does not exist in the selected table."
        )

    observed_headers = (
        [_cell_text(value) for value in raw.iloc[mapping.header_row - 1].tolist()]
        if mapping.header_row is not None
        else [None] * len(raw.columns)
    )
    _validate_references(mapping, observed_headers, len(raw.columns))
    labels = _source_labels(observed_headers, len(raw.columns))
    columns = [
        {"index": index, "header": observed_headers[index], "label": labels[index]}
        for index in range(len(raw.columns))
    ]

    data = raw.iloc[data_start:]
    nonblank_flags = [not _row_is_blank(row) for _, row in data.iterrows()]
    last_nonblank = max(
        (index for index, present in enumerate(nonblank_flags) if present),
        default=-1,
    )
    has_later_data = bool(
        len(raw.index) >= read_limit
        and _has_nonblank_after(
            source, sheet_name=selected_sheet, physical_rows=len(raw.index)
        )
    )
    if last_nonblank >= max_rows or has_later_data:
        raise ValueError(
            f"Input contains more than {max_rows:,} physical data rows; "
            f"limit is {max_rows:,}."
        )

    results: list[tuple[int, dict[str, object], dict[str, str]]] = []
    blank_rows = 0
    missing_firm_sample: list[dict[str, object]] = []
    missing_firm_count = 0
    for raw_index, row in data.iloc[: last_nonblank + 1].iterrows():
        physical_row = int(raw_index) + 1
        if _row_is_blank(row):
            blank_rows += 1
            continue
        source_values = {
            label: _source_value(row.iloc[index])
            for index, label in enumerate(labels)
        }
        mapped = {
            field: _cell_text(row.iloc[reference.index]) if reference else ""
            for field in MAPPED_FIELDS
            for reference in (getattr(mapping, field),)
        }
        if _requires_missing_firm_confirmation(mapped):
            missing_firm_count += 1
            if len(missing_firm_sample) < missing_firm_sample_limit:
                missing_firm_sample.append(
                    {
                        "source_row_number": physical_row,
                        "name": _display_name(mapped),
                        "city": mapped.get("city", ""),
                        "state": mapped.get("state", ""),
                    }
                )
        results.append((physical_row, source_values, mapped))

    if not results:
        raise ValueError("The selected table contains no nonblank data rows.")

    mapping_fingerprint = input_mapping_fingerprint(
        source_sha256=source.sha256,
        mapping=mapping,
    )
    summary = InputSummary(
        data_row_count=len(results),
        blank_row_count=blank_rows,
        preamble_row_count=(mapping.header_row - 1)
        if mapping.header_row is not None
        else 0,
        firm_column_missing=mapping.firm_name is None,
        missing_firm_row_count=missing_firm_count,
        missing_firm_confirmation_required=missing_firm_count > 0,
    )
    return LoadedInput(
        rows=results,
        selected_sheet=selected_sheet,
        columns=columns,
        source_sha256=source.sha256,
        mapping_fingerprint=mapping_fingerprint,
        summary=summary,
        missing_firm_sample=missing_firm_sample,
    )


def load_input(
    source: InMemoryFile, mapping: InputMapping, *, max_rows: int
) -> list[tuple[int, dict[str, object], dict[str, str]]]:
    return validate_and_load_input(source, mapping, max_rows=max_rows).rows


def input_mapping_fingerprint(*, source_sha256: str, mapping: InputMapping) -> str:
    canonical = json.dumps(
        mapping.model_dump(mode="json"),
        ensure_ascii=False,
        separators=(",", ":"),
        sort_keys=True,
    )
    return hashlib.sha256(f"{source_sha256}:{canonical}".encode()).hexdigest()


def _selected_sheet(source: InMemoryFile, mapping: InputMapping) -> str | None:
    if source.suffix == ".csv":
        if mapping.sheet_name is not None:
            raise ValueError("CSV mappings cannot select a worksheet.")
        return None
    try:
        with pd.ExcelFile(source.open()) as excel:
            if not excel.sheet_names:
                raise ValueError("The Excel workbook contains no worksheets.")
            if mapping.sheet_name is None:
                raise ValueError("Excel mappings must select an exact worksheet.")
            selected = mapping.sheet_name
            if selected not in excel.sheet_names:
                raise ValueError(f"Worksheet {selected!r} does not exist.")
            return selected
    except (OSError, ValueError, BadZipFile, InvalidFileException) as exc:
        if isinstance(exc, ValueError) and str(exc).startswith(
            ("Worksheet", "Excel mappings")
        ):
            raise
        raise InvalidUploadError(
            f"The input table could not be read: {exc}"
        ) from exc


def _read_raw(
    source: InMemoryFile, *, sheet_name: str | None, nrows: int
) -> pd.DataFrame:
    try:
        if source.suffix == ".csv":
            text = source.content.decode("utf-8-sig")
            return pd.read_csv(
                StringIO(text),
                header=None,
                dtype=str,
                keep_default_na=False,
                skip_blank_lines=False,
                nrows=nrows,
                sep=_csv_separator(text),
            )
        return pd.read_excel(
            source.open(),
            sheet_name=sheet_name or 0,
            header=None,
            dtype=str,
            keep_default_na=False,
            nrows=nrows,
        )
    except (OSError, UnicodeError, ValueError, BadZipFile, InvalidFileException) as exc:
        raise InvalidUploadError(
            f"The input table could not be read: {exc}"
        ) from exc


def _has_nonblank_after(
    source: InMemoryFile, *, sheet_name: str | None, physical_rows: int
) -> bool:
    if source.suffix == ".csv":
        text = source.content.decode("utf-8-sig")
        reader = csv.reader(StringIO(text), delimiter=_csv_separator(text))
        return any(
            any(str(value or "").strip() for value in row)
            for index, row in enumerate(reader)
            if index >= physical_rows
        )
    workbook = load_workbook(source.open(), read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
        return any(
            any(str(value or "").strip() for value in row)
            for index, row in enumerate(sheet.iter_rows(values_only=True))
            if index >= physical_rows
        )
    finally:
        workbook.close()


def _validate_references(
    mapping: InputMapping,
    observed_headers: list[str | None],
    column_count: int,
) -> None:
    for field, reference in mapping.field_refs().items():
        if reference.index >= column_count:
            raise ValueError(
                f"Mapped field {field!r} references missing column index "
                f"{reference.index}."
            )
        observed = observed_headers[reference.index]
        if observed != reference.header:
            raise ValueError(
                f"Column mapping no longer matches index {reference.index}: "
                f"expected {reference.header!r}, observed {observed!r}."
            )


def _source_labels(
    observed_headers: list[str | None], column_count: int
) -> list[str]:
    used: dict[str, int] = {}
    result = []
    for index in range(column_count):
        header = observed_headers[index]
        base = (
            header.strip()
            if isinstance(header, str) and header.strip()
            else f"Column {get_column_letter(index + 1)}"
        )
        used[base] = used.get(base, 0) + 1
        result.append(base if used[base] == 1 else f"{base} [{used[base]}]")
    return result


def _requires_missing_firm_confirmation(mapped: dict[str, str]) -> bool:
    return bool(
        _usable_name(mapped)
        and not norm.firm(mapped.get("firm_name"))
        and not norm.crd(mapped.get("crd_number"))
        and not norm.email(mapped.get("email"))
    )


def _usable_name(mapped: dict[str, str]) -> str:
    full = norm.person_name(mapped.get("full_name"))
    if len(full.split()) >= 2:
        return full
    first = norm.first_name(mapped.get("first_name"))
    last = norm.person_name(mapped.get("last_name"))
    return f"{first} {last}".strip() if first and last else ""


def _display_name(mapped: dict[str, str]) -> str:
    return str(mapped.get("full_name") or "").strip() or " ".join(
        value
        for value in (
            str(mapped.get("first_name") or "").strip(),
            str(mapped.get("last_name") or "").strip(),
        )
        if value
    )


def _row_is_blank(row: pd.Series) -> bool:
    return all(not _cell_text(value) for value in row.tolist())


def _cell_text(value: object) -> str:
    return "" if pd.isna(value) else str(value).strip()


def _source_value(value: object) -> object:
    return "" if pd.isna(value) else value


def _csv_separator(text: str) -> str:
    try:
        return csv.Sniffer().sniff(text[:65_536], delimiters=",\t;|").delimiter
    except csv.Error:
        return ","
