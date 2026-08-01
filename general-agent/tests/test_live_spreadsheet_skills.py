from __future__ import annotations

import csv
import os
from pathlib import Path

import pytest
from langgraph.checkpoint.memory import InMemorySaver
from openpyxl import Workbook, load_workbook

from general_agent.agent import build_agent
from general_agent.config import load_settings
from general_agent.execution import CancellableLocalShellBackend
from general_agent.workspace import Workspace


@pytest.mark.live
@pytest.mark.asyncio
async def test_live_agent_processes_csv_and_xlsx_with_xlsx_skill() -> None:
    if os.getenv("GENERAL_AGENT_LIVE_TEST") != "1":
        pytest.skip("Set GENERAL_AGENT_LIVE_TEST=1 to make provider model calls.")

    settings = load_settings()
    workspace = Workspace(settings.workspace_root, settings.data_root)
    corp_id = "A123456"
    chat_id = "live-spreadsheet-skill"
    run_id = "live-spreadsheet-skill-run"
    chat = workspace.ensure_chat(corp_id, chat_id)
    uploads = chat / "uploads"
    uploads.mkdir(exist_ok=True)

    (uploads / "sales.csv").write_text(
        "invoice_id,date,revenue,cost\n"
        "0001,2026-07-01,100,60\n"
        "0002,2026-07-02,250,175\n",
        encoding="utf-8",
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Budget"
    sheet.append(["Category", "Budget", "Actual", "Variance"])
    sheet.append(["Software", 1000, 900, None])
    sheet.append(["Travel", 500, 650, None])
    sheet.append(["Total", "=SUM(B2:B3)", "=SUM(C2:C3)", None])
    workbook.save(uploads / "budget.xlsx")

    backend = CancellableLocalShellBackend(
        settings.workspace_root,
        package_root=settings.package_root,
        temp_root=settings.temp_root,
        timeout=settings.command_timeout_seconds,
        max_output_bytes=settings.max_command_output_bytes,
        max_file_read_chars=settings.max_file_read_chars,
    )
    graph = build_agent(
        settings,
        workspace=workspace,
        backend=backend,
        checkpointer=InMemorySaver(),
    )
    prompt = """
Use the xlsx skill and its verification workflow for both tasks. Do not merely
describe the files; create the exact outputs requested.

1. Read /uploads/sales.csv and create /sales_cleaned.csv. Preserve invoice_id
as text including leading zeroes, preserve the ISO dates, and add margin_pct as
(revenue-cost)/revenue using decimal values (0.4 means 40%).
2. Edit /uploads/budget.xlsx into /budget_completed.xlsx without overwriting the
input. Preserve the existing formulas and formatting conventions. Add formulas
in D2:D4 for Actual minus Budget, apply a currency number format to the numeric
columns, recalculate with the bundled recalc.py, and verify there are no formula
errors and that cached values are present.

Finish with a concise summary naming both output files.
""".strip()

    async with backend.run_scope(run_id, corp_id, chat_id):
        result = await graph.ainvoke(
            {"messages": [{"role": "user", "content": prompt}]},
            config={"configurable": {"thread_id": run_id}},
        )
    assert result.get("messages")

    cleaned = chat / "sales_cleaned.csv"
    completed = chat / "budget_completed.xlsx"
    assert cleaned.is_file()
    assert completed.is_file()

    with cleaned.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    assert [row["invoice_id"] for row in rows] == ["0001", "0002"]
    assert float(rows[0]["margin_pct"]) == pytest.approx(0.4)
    assert float(rows[1]["margin_pct"]) == pytest.approx(0.3)

    formulas = load_workbook(completed, data_only=False)
    cached = load_workbook(completed, data_only=True)
    budget = formulas["Budget"]
    values = cached["Budget"]
    assert budget["B4"].value == "=SUM(B2:B3)"
    assert budget["C4"].value == "=SUM(C2:C3)"
    assert budget["D2"].value.startswith("=")
    assert budget["D3"].value.startswith("=")
    assert budget["D4"].value.startswith("=")
    assert values["D2"].value == -100
    assert values["D3"].value == 150
    assert values["D4"].value == 50
    formulas.close()
    cached.close()
