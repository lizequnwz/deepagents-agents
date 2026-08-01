from __future__ import annotations

import csv
import json
import os
import re
import subprocess
import sys
from pathlib import Path

import pandas as pd
import pytest
import yaml
from openpyxl import Workbook, load_workbook
from reportlab.pdfgen import canvas

from general_agent.workspace import Workspace


EXPECTED_SKILLS = {
    "docx",
    "frontend-design",
    "pdf",
    "pptx",
    "skill-creator",
    "theme-factory",
    "xlsx",
}


def _skill_env(settings, workspace: Workspace, chat_id: str) -> dict[str, str]:
    chat = workspace.ensure_chat("A123456", chat_id)
    temp = workspace.temp_root("A123456") / "skill-test"
    temp.mkdir(parents=True, exist_ok=True)
    return {
        **os.environ,
        "GENERAL_AGENT_CHAT_DIR": str(chat),
        "GENERAL_AGENT_SHARED_DIR": str(workspace.shared_root("A123456")),
        "GENERAL_AGENT_TEMP_DIR": str(temp),
    }


def _run_manifest(script: Path, source: Path, env: dict[str, str]) -> dict:
    result = subprocess.run(
        [sys.executable, str(script), str(source)],
        cwd=source.parent,
        env=env,
        text=True,
        capture_output=True,
        check=True,
    )
    manifest = json.loads(result.stdout)
    output = Path(env["GENERAL_AGENT_TEMP_DIR"]) / Path(
        manifest["virtual_path"]
    ).name
    assert output.is_file()
    assert output.stat().st_size == manifest["characters"]
    return {**manifest, "output": output}


def test_main_skill_tree_uses_canonical_packages() -> None:
    skills = Path(__file__).resolve().parents[1] / "skills"
    actual = {path.name for path in skills.iterdir() if (path / "SKILL.md").is_file()}
    assert actual == EXPECTED_SKILLS

    for name in sorted(actual):
        content = (skills / name / "SKILL.md").read_text(encoding="utf-8")
        match = re.match(r"^---\n(.*?)\n---", content, re.DOTALL)
        assert match, f"{name} has invalid frontmatter delimiters"
        frontmatter = yaml.safe_load(match.group(1))
        assert frontmatter["name"] == name
        assert isinstance(frontmatter["description"], str)
        assert frontmatter["description"].strip()

    assert (skills / "frontend-design/LICENSE.txt").is_file()
    assert (skills / "theme-factory/LICENSE.txt").is_file()
    assert not list(skills.rglob(".DS_Store"))


def test_skill_markdown_has_no_broken_relative_links() -> None:
    skills = Path(__file__).resolve().parents[1] / "skills"
    link_pattern = re.compile(r"\[[^]]*]\(([^)]+)\)")
    for skill_md in skills.glob("*/SKILL.md"):
        for target in link_pattern.findall(skill_md.read_text(encoding="utf-8")):
            if target.startswith(("#", "http://", "https://")):
                continue
            relative = target.split("#", 1)[0]
            assert (skill_md.parent / relative).exists(), (
                f"{skill_md.relative_to(skills)} links to missing {target}"
            )


def test_prepare_directories_removes_retired_installed_skills(settings) -> None:
    source = settings.skills_source_root
    (source / "xlsx").mkdir(parents=True)
    (source / "xlsx/SKILL.md").write_text(
        "---\nname: xlsx\ndescription: test\n---\n", encoding="utf-8"
    )
    stale = settings.installed_skills_root / "spreadsheets"
    stale.mkdir(parents=True)
    (stale / "SKILL.md").write_text(
        "---\nname: spreadsheets\ndescription: stale\n---\n", encoding="utf-8"
    )

    settings.prepare_directories()

    assert (settings.installed_skills_root / "xlsx/SKILL.md").is_file()
    assert not stale.exists()


def test_pdf_skill_keeps_bounded_initial_inspection(settings) -> None:
    workspace = Workspace(settings.workspace_root, settings.data_root)
    chat = workspace.ensure_chat("A123456", "pdf-skill")
    env = _skill_env(settings, workspace, "pdf-skill")
    pdf = chat / "sample.pdf"
    drawing = canvas.Canvas(str(pdf))
    drawing.drawString(72, 720, "PDF bounded extraction")
    drawing.save()

    result = _run_manifest(
        Path(__file__).resolve().parents[1] / "skills/pdf/scripts/extract_pdf.py",
        pdf,
        env,
    )

    assert result["page_count"] == 1
    assert "PDF bounded extraction" in result["output"].read_text()


def test_xlsx_workflow_processes_csv_and_recalculates_formulas(tmp_path: Path) -> None:
    source_csv = tmp_path / "sales.csv"
    source_csv.write_text(
        "invoice_id,revenue,cost\n0001,100,60\n0002,250,175\n",
        encoding="utf-8",
    )
    frame = pd.read_csv(source_csv, dtype={"invoice_id": "string"})
    frame["margin_pct"] = (frame["revenue"] - frame["cost"]) / frame["revenue"]
    cleaned_csv = tmp_path / "sales-cleaned.csv"
    frame.to_csv(cleaned_csv, index=False)

    with cleaned_csv.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    assert rows[0]["invoice_id"] == "0001"
    assert float(rows[0]["margin_pct"]) == pytest.approx(0.4)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales"
    sheet.append(["Invoice ID", "Revenue", "Cost", "Margin"])
    for row_index, row in enumerate(rows, start=2):
        sheet.cell(row_index, 1, row["invoice_id"]).number_format = "@"
        sheet.cell(row_index, 2, float(row["revenue"])).number_format = "$#,##0.00"
        sheet.cell(row_index, 3, float(row["cost"])).number_format = "$#,##0.00"
        sheet.cell(row_index, 4, f"=(B{row_index}-C{row_index})/B{row_index}")
        sheet.cell(row_index, 4).number_format = "0.0%"
    output = tmp_path / "sales.xlsx"
    workbook.save(output)

    skills = Path(__file__).resolve().parents[1] / "skills"
    result = subprocess.run(
        [sys.executable, str(skills / "xlsx/scripts/recalc.py"), str(output), "60"],
        text=True,
        capture_output=True,
    )
    payload = json.loads(result.stdout)
    if "soffice not found" in payload.get("error", ""):
        pytest.skip("LibreOffice is required for XLSX recalculation")
    assert result.returncode == 0, result.stderr or result.stdout
    assert payload == {
        "status": "success",
        "total_errors": 0,
        "error_summary": {},
        "total_formulas": 2,
    }

    formulas = load_workbook(output, data_only=False)
    values = load_workbook(output, data_only=True)
    assert formulas["Sales"]["D2"].value == "=(B2-C2)/B2"
    assert values["Sales"]["D2"].value == pytest.approx(0.4)
    assert values["Sales"]["A2"].value == "0001"
    formulas.close()
    values.close()
